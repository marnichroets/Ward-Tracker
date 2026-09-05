"""Official Capture Workspace (Phase 4) — completely separate from
smartsheet_reporting.py's classification/bucketing logic. This module never
imports CANONICAL_ACTIVITY_CATEGORY, classify_activity_text, or any
SmartSheet category concept, and never renames or reclassifies the original
Ward Tracker `type`/`type_display` fields those depend on.

Two distinct concepts live side by side on the same entry document:
1. The original Ward Tracker activity type (`type`/`type_display`) — never
   touched by anything here.
2. The official Campaign Manager activity type (`official_activity_type`) —
   either a confident automatic suggestion (computed fresh on every read,
   never persisted just because it exists) or an explicit coordinator
   override (the only thing ever written to the database field).

Two separate lists exist here, deliberately not merged:
- `OFFICIAL_ACTIVITY_TYPES` — the FULL known official Campaign Manager type
  list (46 types, supplied directly by the coordinator). This is what the
  admin dropdown offers and what a manual override is validated against —
  a coordinator can confirm ANY Ward Tracker activity, confidently-mapped
  or not, as any type on this full list.
- `_CONFIDENT_MAP_RAW` — a deliberately narrower subset used ONLY to compute
  an automatic *suggestion*. Never guesses: a Ward Tracker activity not in
  this table gets no suggestion at all ("Official type needs confirmation"),
  regardless of how large the full list above is.
"""

import io
import re
from typing import Iterable, Optional

# openpyxl (and smartsheet_reporting, which also needs it for its own
# exports) is only actually required by official_capture_xlsx_bytes below —
# imported lazily there, not at module level, so every other function in
# this module (the mapping table, capture-status helpers, augment/filter/
# sort) stays importable and testable in an environment that has no Excel
# library installed at all.

AWAITING_CAPTURE = "awaiting_capture"
CAPTURED = "captured"
CAPTURE_STATUSES = (AWAITING_CAPTURE, CAPTURED)
CAPTURE_STATUS_LABELS = {AWAITING_CAPTURE: "Awaiting Capture", CAPTURED: "Captured"}

NEEDS_CONFIRMATION_LABEL = "Official type needs confirmation"
# Sentinel accepted by the official_activity_type filter/query param to mean
# "resolved type is empty" — distinct from any real official type string.
NEEDS_CONFIRMATION_FILTER_VALUE = "Needs confirmation"

# Confident Ward Tracker -> official Campaign Manager mappings, audited per
# CAMPAIGNS.md §4. Keyed off the raw candidate-typed activity text
# (type_display, falling back to type) — never off canonical_activity or
# smartsheet_category — so a future SmartSheet classification change can
# never silently change an official-type suggestion, or vice versa.
# Anything not listed here is deliberately left unmapped: this table never
# guesses.
_CONFIDENT_MAP_RAW = {
    "Door to Door": "In-person Canvassing / Door-to-door",
    "Info Table": "Info Table",
    "Info table : canvassing": "Info Table",
    "Telecanvassing": "Tele Canvassing",
    "House Meeting": "House meeting",
    "Public Meeting": "Public meeting",
    "Clean up": "Clean-up Event",
    "Oversight": "Oversight Visit",
    "Stakeholder meeting": "Stakeholder Meeting",
    "Hoot or Blue wave": "Blue Wave / Robot blitz",
    "Blue Wave": "Blue Wave / Robot blitz",
    "Motorcade": "Cavalcade / Carcade / Motorcade",
    "March": "March",
    "Picket": "Protest / Picket",
    "Rally": "Rally",
    "Religious Forum Address": "Religious Forum Address",
    "Poster fighting": "Poster fighting",
    "Leaflet Distribution": "Leaflet distribution",
    "Care Event": "Care event (Oppit)",
}


def _normalise(value: object) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[-_/:]+", " ", text)
    text = re.sub(r"[^a-z0-9\s]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


_SUGGESTED_MAP = {_normalise(k): v for k, v in _CONFIDENT_MAP_RAW.items()}

# The FULL known official Campaign Manager activity type list — supplied
# directly by the coordinator, not derived from the confident-mapping table
# above. This is what the admin dropdown offers and what a manual
# official_activity_type override is validated against, independent of
# whether that particular type happens to also be a confident auto-suggestion
# target. Ambiguous/unmapped Ward Tracker activities get NO automatic
# suggestion (see suggested_official_type below) but the coordinator can
# still manually confirm any of these 46 types for them.
OFFICIAL_ACTIVITY_TYPES = sorted([
    "Billboard",
    "Blue Wave / Robot blitz",
    "Bulletin Boards",
    "Care collection drive",
    "Care event (Oppit)",
    "Cavalcade / Carcade / Motorcade",
    "Clean-up Event",
    "Community Crime Patrol",
    "Community Sporting Event",
    "Delivery failure site visit",
    "Email send",
    "Federal Leader Event",
    "Front of House",
    "House meeting",
    "In-person Canvassing / Door-to-door",
    "Info Table",
    "Leaflet distribution",
    "Loudhailing",
    "March",
    "Microtargeting - In-Person Survey / Door-to-door",
    "Newspaper advert",
    "NGO/NPO Assistance",
    "Oversight Visit",
    "Podcast interview",
    "Poster fighting",
    "Press conference",
    "Press statement",
    "Protest / Picket",
    "Public meeting",
    "Queue Assistance",
    "Radio interview",
    "Rally",
    "Registration Surgery",
    "Religious Forum Address",
    "Roadmarkings",
    "Robocalls",
    "Self canvass(es)",
    "SMS send",
    "Social media advert",
    "Social media post",
    "Social media promoted post",
    "Sound truck",
    "Stakeholder Meeting",
    "Tele Canvassing",
    "Television interview",
    "WhatsApp/Telegram",
])

# Living documentation + a fail-loud guard: every confident auto-suggestion
# target must actually exist on the full list, or a suggestion could point
# at a type the admin dropdown/validator would then reject.
assert set(_CONFIDENT_MAP_RAW.values()) <= set(OFFICIAL_ACTIVITY_TYPES), (
    "a confident-mapping target is missing from OFFICIAL_ACTIVITY_TYPES"
)


def entry_activity_text(doc: dict) -> str:
    return str(doc.get("type_display") or doc.get("type") or "").strip()


def suggested_official_type(doc: dict) -> Optional[str]:
    """Confident, automatic suggestion only — never guesses. Returns None
    (displayed as "Official type needs confirmation") for any Ward Tracker
    activity text not in the audited confident-mapping table above."""
    return _SUGGESTED_MAP.get(_normalise(entry_activity_text(doc)))


def resolve_official_activity_type(doc: dict) -> Optional[str]:
    """The value to actually display/export: a coordinator-confirmed
    override always wins; otherwise fall back to the automatic suggestion."""
    stored = doc.get("official_activity_type")
    if stored:
        return stored
    return suggested_official_type(doc)


def validate_official_activity_type(value: str) -> str:
    value = (value or "").strip()
    if value not in OFFICIAL_ACTIVITY_TYPES:
        raise ValueError("Unknown official activity type")
    return value


def resolve_capture_status(doc: dict) -> str:
    """Historical documents with no capture_status field (or an unexpected
    value) safely resolve to awaiting_capture at read time — this never
    writes anything back to the document."""
    status = doc.get("capture_status")
    return status if status in CAPTURE_STATUSES else AWAITING_CAPTURE


def validate_capture_status(value: str) -> str:
    value = (value or "").strip()
    if value not in CAPTURE_STATUSES:
        raise ValueError("Invalid capture status")
    return value


def augment_entry(doc: dict, campaign_name: Optional[str]) -> dict:
    """Build one Official Capture Workspace row from an already
    response-shaped entry dict (see main.entry_for_response). Never mutates
    or persists anything — purely a display/reporting projection."""
    official_override = doc.get("official_activity_type")
    suggested = suggested_official_type(doc)
    resolved_type = official_override or suggested
    if official_override:
        type_source = "confirmed"
    elif suggested:
        type_source = "suggested"
    else:
        type_source = "unmapped"

    return {
        "id": doc.get("id"),
        "person_id": doc.get("person_id"),
        "activity_date": doc.get("activity_date") or "",
        "start_time": doc.get("start_time"),
        "end_time": doc.get("end_time"),
        "name": doc.get("name", ""),
        "ward": doc.get("ward", ""),
        "campaign_id": doc.get("campaign_id"),
        "campaign_name": campaign_name,
        "type_display": doc.get("type_display") or doc.get("type") or "",
        "venue": doc.get("venue"),
        "official_activity_type": resolved_type,
        "official_activity_type_source": type_source,
        "capture_status": resolve_capture_status(doc),
        "captured_at": doc.get("captured_at"),
    }


def compute_counts(entries: Iterable[dict]) -> dict:
    entries = list(entries)
    awaiting = sum(1 for e in entries if e.get("capture_status") == AWAITING_CAPTURE)
    captured = sum(1 for e in entries if e.get("capture_status") == CAPTURED)
    return {"awaiting_capture": awaiting, "captured": captured, "total": len(entries)}


def filter_entries(
    entries: Iterable[dict],
    *,
    status: Optional[str] = None,
    person_id: Optional[str] = None,
    campaign_id: Optional[str] = None,
    official_activity_type: Optional[str] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
) -> list[dict]:
    def keep(e: dict) -> bool:
        if status and status != "all" and e.get("capture_status") != status:
            return False
        if person_id and e.get("person_id") != person_id:
            return False
        if campaign_id and (e.get("campaign_id") or "") != campaign_id:
            return False
        if official_activity_type:
            if official_activity_type == NEEDS_CONFIRMATION_FILTER_VALUE:
                if e.get("official_activity_type"):
                    return False
            elif e.get("official_activity_type") != official_activity_type:
                return False
        activity_date = e.get("activity_date") or ""
        if date_from and activity_date < date_from:
            return False
        if date_to and activity_date > date_to:
            return False
        return True

    return [e for e in entries if keep(e)]


def sort_oldest_first(entries: Iterable[dict]) -> list[dict]:
    return sorted(
        entries,
        key=lambda e: (e.get("activity_date") or "", e.get("start_time") or "", e.get("name") or ""),
    )


OFFICIAL_CAPTURE_HEADERS = [
    "DATE",
    "START TIME",
    "END TIME",
    "CANDIDATE",
    "WARD",
    "CAMPAIGN",
    "WARD TRACKER ACTIVITY",
    "OFFICIAL ACTIVITY TYPE",
    "LOCATION / VENUE",
    "CAPTURE STATUS",
    "CAPTURED AT",
]


def official_capture_xlsx_bytes(entries: Iterable[dict]) -> bytes:
    """Plain, paste-friendly worksheet: header row 1, data from row 2, no
    title/metadata rows, no merged cells, no formulas, no macros — one value
    per cell. Candidate-controlled free text (name/ward/campaign name/
    activity/venue) is passed through the existing spreadsheet-injection
    guard, reused as-is rather than duplicated."""
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
    from smartsheet_reporting import spreadsheet_safe_text

    wb = Workbook()
    ws = wb.active
    ws.title = "Official Capture"
    ws.append(OFFICIAL_CAPTURE_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    rows = []
    for e in entries:
        rows.append([
            e.get("activity_date") or "",
            e.get("start_time") or "",
            e.get("end_time") or "",
            spreadsheet_safe_text(e.get("name") or ""),
            spreadsheet_safe_text(e.get("ward") or ""),
            spreadsheet_safe_text(e.get("campaign_name") or "—"),
            spreadsheet_safe_text(e.get("type_display") or ""),
            e.get("official_activity_type") or NEEDS_CONFIRMATION_LABEL,
            spreadsheet_safe_text(e.get("venue") or ""),
            CAPTURE_STATUS_LABELS.get(e.get("capture_status"), CAPTURE_STATUS_LABELS[AWAITING_CAPTURE]),
            e.get("captured_at") or "",
        ])
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(OFFICIAL_CAPTURE_HEADERS, start=1):
        max_len = len(header)
        for row in rows:
            value = row[col_idx - 1]
            if value:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
