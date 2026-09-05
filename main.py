# trigger redeploy
import os
import re
import io
import csv
from collections import Counter
from datetime import date, datetime, timedelta, timezone
from typing import Optional, List

from fastapi import FastAPI, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
from jose import jwt, JWTError
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.drawing.image import Image as XLImage
import certifi
import ssl

from week_dates import (
    DAY_LABELS,
    DAY_OFFSET,
    DAY_ORDER,
    MONTHS,
    activity_date_for_day,
    activity_date_for_day_date,
    current_week_key,
    format_week_label,
    normalise_new_activity_date,
    sast_today,
    validate_campaign_date_range,
    validate_candidate_week_key,
)
from smartsheet_reporting import (
    CANVASSING,
    DEFAULT_CONSTITUENCY,
    PRESENCE,
    PUBLIC_STREET_MEETING,
    REVIEWABLE_CATEGORIES,
    is_new_other_submission,
    review_entries,
    reporting_metadata_for_submission,
    smartsheet_csv_bytes,
    smartsheet_workbook_all_categories_bytes,
    smartsheet_xlsx_bytes,
    summarize_smartsheet_entries,
    validate_time_range,
    normalise_venue,
)


# Atlas on this host rejects TLS 1.3 (TLSV1_ALERT_INTERNAL_ERROR); cap at TLS 1.2.
_original_create_default_context = ssl.create_default_context


def _create_default_context_tls12(*args, **kwargs):
    context = _original_create_default_context(*args, **kwargs)
    context.maximum_version = ssl.TLSVersion.TLSv1_2
    return context


ssl.create_default_context = _create_default_context_tls12

# ---------- Config ----------
MONGO_URI = os.environ["MONGO_URI"]                     # required - set in Railway
DB_NAME = os.environ.get("DB_NAME", "ward_tracker")
ADMIN_PIN = os.environ["ADMIN_PIN"]                      # required - your coordinator PIN
JWT_SECRET = os.environ["JWT_SECRET"]                    # required - long random string
JWT_ALGO = "HS256"
JWT_EXPIRE_DAYS = 30
FRONTEND_ORIGIN = os.environ.get("FRONTEND_ORIGIN", "*")  # set to your Vercel URL once deployed
LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "static", "logo.png")
CONSTITUENCY = os.environ.get("CONSTITUENCY", DEFAULT_CONSTITUENCY)

app = FastAPI(title="Ntsikana Ward Tracker API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[FRONTEND_ORIGIN] if FRONTEND_ORIGIN != "*" else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

client = AsyncIOMotorClient(MONGO_URI, tlsCAFile=certifi.where())
db = client[DB_NAME]
entries_col = db["entries"]
roster_col = db["roster"]
campaigns_col = db["campaigns"]


@app.on_event("startup")
async def ensure_indexes():
    await entries_col.create_index([("person_id", 1), ("week_key", 1)])
    await roster_col.create_index("name_slug", unique=True)
    # Additive only: existing entries documents have no campaign_id field at
    # all, which Mongo indexes identically to campaign_id: null — no backfill
    # needed for this index to be useful once campaign-linked activities exist.
    await entries_col.create_index("campaign_id")
    await campaigns_col.create_index([("person_id", 1), ("start_date", -1)])


# ---------- Helpers ----------
def slugify(s: str) -> str:
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")


def oid_str(doc: dict) -> dict:
    doc = dict(doc)
    doc["id"] = str(doc.pop("_id"))
    return doc


def enrich_entry(doc: dict) -> dict:
    doc = dict(doc)
    week_key = doc.get("week_key")
    day = doc.get("day")
    if week_key:
        doc["week_label"] = format_week_label(week_key)
    if week_key and day in DAY_OFFSET and not doc.get("activity_date"):
        doc["activity_date"] = activity_date_for_day(week_key, day)
    return doc


def entry_for_response(doc: dict) -> dict:
    return enrich_entry(oid_str(doc))


def normalize_name_words(name: str) -> set:
    # Strips titles like "(CLLR)" so a name typed without middle names/suffixes
    # still matches its fuller roster form.
    name = re.sub(r"\([^)]*\)", "", name)
    name = re.sub(r"[^a-z0-9\s]", "", name.lower())
    return set(w for w in name.split() if w)


def names_match(a: str, b: str) -> bool:
    words_a, words_b = normalize_name_words(a), normalize_name_words(b)
    if not words_a or not words_b:
        return False
    return words_a <= words_b or words_b <= words_a


async def resolve_and_canonicalize_person(body: "EntryIn") -> None:
    """Validate the submitted candidate against the official roster and replace
    whatever name/ward/person_id the client sent with the roster's canonical
    values. This is the backend enforcement point: it runs for every entry
    create/update regardless of whether the request came through the UI, so a
    direct API call cannot invent a new person or attach itself to the wrong
    roster identity by spelling/case/abbreviation variation."""
    roster_person = await roster_col.find_one({"name_slug": slugify(body.name)})
    if not roster_person:
        raise HTTPException(400, "Please select your name from the roster.")
    body.name = roster_person["name"]
    body.ward = roster_person.get("ward", "")
    body.person_id = roster_person["name_slug"]


async def resolve_campaign_person_id(person_id: str) -> str:
    """Same roster-only identity enforcement as entries: a campaign may only
    belong to an actual roster person. `person_id` here is whatever identity
    string the client is currently using (today's frontend always sends the
    slug of the selected roster person's canonical name, which already
    equals that roster entry's name_slug) — re-slugifying it is a no-op for
    an already-canonical slug and correctly resolves a raw name too. The
    client-sent value is never trusted directly; only the roster's own
    name_slug is stored, exactly as resolve_and_canonicalize_person does for
    entries."""
    roster_person = await roster_col.find_one({"name_slug": slugify(person_id)})
    if not roster_person:
        raise HTTPException(400, "Please select your name from the roster.")
    return roster_person["name_slug"]


def campaign_doc_from_body(body: "CampaignIn") -> dict:
    """Deliberately excludes person_id: ownership is resolved and applied
    separately by each route (set once at create, never accepted from an
    update body afterwards), so a mutable field never sneaks back in here."""
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(400, "Campaign name is required")
    try:
        start, end = validate_campaign_date_range(body.start_date, body.end_date)
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    return {
        "name": name,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
    }


def derive_campaign_status(doc: dict, now: Optional[datetime] = None) -> str:
    """Planned/active/completed are pure functions of today vs the stored
    dates and are deliberately NOT persisted, so they can never drift out of
    sync with the calendar. archived_at is the only stored, manual state."""
    if doc.get("archived_at"):
        return "archived"
    today = sast_today(now)
    start = date.fromisoformat(doc["start_date"])
    end = date.fromisoformat(doc["end_date"])
    if today < start:
        return "planned"
    if today > end:
        return "completed"
    return "active"


def campaign_for_response(doc: dict) -> dict:
    doc = oid_str(doc)
    doc["status"] = derive_campaign_status(doc)
    return doc


async def require_campaign_owner(campaign_id: str, person_id: str) -> tuple[dict, str]:
    """Resolve the caller's claimed identity against the roster (rejecting an
    unknown identity exactly like create_campaign does), then load the
    campaign only if it belongs to that resolved identity. A campaign that
    exists but belongs to someone else is deliberately indistinguishable
    from one that doesn't exist (404) — the same convention update_entry/
    delete_entry already use for person-scoped entries, so a wrong guess at
    someone else's campaign_id never confirms it exists or reveals its
    owner. There is no admin bypass here; this is candidate-facing only."""
    owner_person_id = await resolve_campaign_person_id(person_id)
    doc = await campaigns_col.find_one({"_id": ObjectId(campaign_id), "person_id": owner_person_id})
    if not doc:
        raise HTTPException(404, "Campaign not found")
    return doc, owner_person_id


def entry_doc_from_body(body: "EntryIn", existing_doc: Optional[dict] = None) -> dict:
    doc = body.model_dump()
    try:
        validate_candidate_week_key(doc["week_key"])
        doc["week_label"] = format_week_label(doc["week_key"])
        doc["activity_date"] = normalise_new_activity_date(
            doc["week_key"], doc["day"], doc.get("activity_date")
        )
        doc["start_time"], doc["end_time"] = validate_time_range(
            doc.get("start_time"), doc.get("end_time")
        )
        doc["venue"] = normalise_venue(doc.get("venue"))
        if is_new_other_submission(doc):
            other_text = (doc.get("type_display") or "").strip()
            if not other_text:
                raise ValueError("Other activity text is required")
            doc["type_display"] = other_text
    except ValueError as exc:
        raise HTTPException(400, str(exc))
    if existing_doc and doc.get("notes") is None and existing_doc.get("notes"):
        doc["notes"] = existing_doc["notes"]
    doc.update(reporting_metadata_for_submission(doc, existing_doc))
    return doc


def make_admin_token() -> str:
    payload = {
        "role": "admin",
        "exp": datetime.now(timezone.utc) + timedelta(days=JWT_EXPIRE_DAYS),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)


async def require_admin(authorization: Optional[str] = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Missing admin token")
    token = authorization.split(" ", 1)[1]
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except JWTError:
        raise HTTPException(401, "Invalid or expired admin token")
    if payload.get("role") != "admin":
        raise HTTPException(401, "Invalid admin token")
    return True


# ---------- Schemas ----------
class LoginRequest(BaseModel):
    pin: str


class EntryIn(BaseModel):
    person_id: str
    name: str
    ward: str
    day: str  # mon..sun
    type: str
    type_display: str
    notes: Optional[str] = None
    week_key: str
    week_label: str
    activity_date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    venue: Optional[str] = None


class EntryOut(EntryIn):
    id: str
    submitted_at: str
    canonical_activity: Optional[str] = None
    smartsheet_category: Optional[str] = None
    category_source: Optional[str] = None
    category_reviewed: Optional[bool] = None
    category_reviewed_at: Optional[str] = None
    is_custom_activity: Optional[bool] = None
    # Forward-compatible only: no endpoint sets this yet (that's a later
    # campaign-activity-linking phase). Declaring it now on the admin-facing
    # model means that phase won't need a second response-model change.
    # Absent on every existing historical document; defaults to None.
    campaign_id: Optional[str] = None


# Response shape for the candidate-facing entry endpoints. Deliberately excludes
# SmartSheet/admin-only fields (smartsheet_category, canonical_activity,
# category_source, category_reviewed*) — candidates must never receive that
# terminology, even in a raw network response the UI doesn't render. FastAPI's
# response_model filtering drops any extra fields the handler returns, so this
# is enforced without changing what the handlers themselves compute or store.
class CandidateEntryOut(BaseModel):
    id: str
    day: str
    type: str
    type_display: str
    notes: Optional[str] = None
    week_key: str
    activity_date: Optional[str] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    venue: Optional[str] = None


class RosterIn(BaseModel):
    name: str
    ward: str


class ReassignPersonIn(BaseModel):
    name: str  # must be the exact canonical name of an existing roster entry


class CategoryReviewIn(BaseModel):
    smartsheet_category: str


# A campaign is deliberately minimal: identity + name + a date range. No
# focus type, area, description, target, budget, photo, or documents — those
# are explicitly out of scope for this phase and should not be added back in
# without a fresh decision, not as an incidental side effect of other work.
class CampaignIn(BaseModel):
    person_id: str
    name: str
    start_date: str
    end_date: str


class CampaignOut(BaseModel):
    id: str
    person_id: str
    name: str
    start_date: str
    end_date: str
    status: str
    created_at: str
    archived_at: Optional[str] = None


# ---------- Auth ----------
@app.post("/api/admin/login")
async def admin_login(body: LoginRequest):
    if body.pin != ADMIN_PIN:
        raise HTTPException(401, "Incorrect PIN")
    return {"token": make_admin_token()}


# ---------- Member: entries ----------
@app.get("/api/entries", response_model=List[CandidateEntryOut])
async def list_my_entries(person_id: str, week_key: str):
    cursor = entries_col.find({"person_id": person_id, "week_key": week_key})
    out = []
    async for doc in cursor:
        out.append(entry_for_response(doc))
    return out


@app.post("/api/entries", response_model=CandidateEntryOut)
async def create_entry(body: EntryIn):
    await resolve_and_canonicalize_person(body)
    doc = entry_doc_from_body(body)
    doc["submitted_at"] = datetime.now(timezone.utc).isoformat()
    res = await entries_col.insert_one(doc)
    doc["_id"] = res.inserted_id
    return entry_for_response(doc)


@app.put("/api/entries/{entry_id}", response_model=CandidateEntryOut)
async def update_entry(entry_id: str, body: EntryIn):
    await resolve_and_canonicalize_person(body)
    existing_doc = await entries_col.find_one({"_id": ObjectId(entry_id), "person_id": body.person_id})
    if not existing_doc:
        raise HTTPException(404, "Entry not found")
    doc = entry_doc_from_body(body, existing_doc)
    doc["submitted_at"] = datetime.now(timezone.utc).isoformat()
    result = await entries_col.find_one_and_update(
        {"_id": ObjectId(entry_id), "person_id": body.person_id},
        {"$set": doc},
        return_document=True,
    )
    if not result:
        raise HTTPException(404, "Entry not found")
    return entry_for_response(result)


@app.delete("/api/entries/{entry_id}")
async def delete_entry(entry_id: str, person_id: str):
    result = await entries_col.delete_one({"_id": ObjectId(entry_id), "person_id": person_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Entry not found")
    return {"deleted": True}


# ---------- Member: campaigns ----------
# Phase 1: backend foundation only. Campaigns are a thin, additive container
# — they never rewrite or touch entries documents. No endpoint here creates,
# links, or generates activities; that is a separate later phase.
@app.post("/api/campaigns", response_model=CampaignOut)
async def create_campaign(body: CampaignIn):
    person_id = await resolve_campaign_person_id(body.person_id)
    doc = campaign_doc_from_body(body)
    doc["person_id"] = person_id
    doc["created_at"] = datetime.now(timezone.utc).isoformat()
    doc["archived_at"] = None
    res = await campaigns_col.insert_one(doc)
    doc["_id"] = res.inserted_id
    return campaign_for_response(doc)


@app.get("/api/campaigns", response_model=List[CampaignOut])
async def list_campaigns(person_id: str):
    cursor = campaigns_col.find({"person_id": person_id})
    return [campaign_for_response(doc) async for doc in cursor]


@app.get("/api/campaigns/{campaign_id}", response_model=CampaignOut)
async def get_campaign(campaign_id: str):
    doc = await campaigns_col.find_one({"_id": ObjectId(campaign_id)})
    if not doc:
        raise HTTPException(404, "Campaign not found")
    return campaign_for_response(doc)


# Ownership (person_id) is immutable after creation. body.person_id here is
# only ever used to prove which roster identity is making the request — it
# is never written as a "new" owner. require_campaign_owner already 404s if
# it resolves to anyone other than the campaign's existing owner, and
# campaign_doc_from_body's returned dict has no person_id key at all, so the
# $set below structurally cannot change ownership even if that check were
# ever weakened by a future edit.
@app.put("/api/campaigns/{campaign_id}", response_model=CampaignOut)
async def update_campaign(campaign_id: str, body: CampaignIn):
    _existing, owner_person_id = await require_campaign_owner(campaign_id, body.person_id)
    doc = campaign_doc_from_body(body)
    result = await campaigns_col.find_one_and_update(
        {"_id": ObjectId(campaign_id), "person_id": owner_person_id},
        {"$set": doc},
        return_document=True,
    )
    if not result:
        raise HTTPException(404, "Campaign not found")
    return campaign_for_response(result)


# The normal terminal action. Archiving only ever sets archived_at — it never
# reads, writes, or otherwise touches entries_col, and it must stay that way:
# linked activities remain fully intact, unmodified, and visible everywhere
# they already appear (admin views, exports, candidate history). Scoped to
# the campaign's owner exactly like update_campaign — knowing a campaign_id
# alone is not enough to archive someone else's campaign.
@app.patch("/api/campaigns/{campaign_id}/archive", response_model=CampaignOut)
async def archive_campaign(campaign_id: str, person_id: str):
    _existing, owner_person_id = await require_campaign_owner(campaign_id, person_id)
    result = await campaigns_col.find_one_and_update(
        {"_id": ObjectId(campaign_id), "person_id": owner_person_id},
        {"$set": {"archived_at": datetime.now(timezone.utc).isoformat()}},
        return_document=True,
    )
    if not result:
        raise HTTPException(404, "Campaign not found")
    return campaign_for_response(result)


# Not part of normal candidate UX — a narrow escape hatch for a
# just-created-by-mistake campaign only. Never cascades to entries: any
# campaign with at least one linked activity is rejected with 409, and there
# is deliberately no code path anywhere that deletes or modifies activities
# as a side effect of a campaign being removed. Scoped to the campaign's
# owner exactly like update/archive — ownership is checked (and 404s on
# mismatch) BEFORE the linked-activity check, so a non-owner can't even learn
# whether someone else's campaign has linked activities.
@app.delete("/api/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str, person_id: str):
    await require_campaign_owner(campaign_id, person_id)
    linked = await entries_col.find_one({"campaign_id": campaign_id})
    if linked:
        raise HTTPException(409, "Cannot delete a campaign with linked activities")
    result = await campaigns_col.delete_one({"_id": ObjectId(campaign_id)})
    if result.deleted_count == 0:
        raise HTTPException(404, "Campaign not found")
    return {"deleted": True}


# ---------- Admin: report ----------
@app.get("/api/admin/report")
async def admin_report(week_key: str, _: bool = Depends(require_admin)):
    cursor = entries_col.find({"week_key": week_key})
    entries = [entry_for_response(doc) async for doc in cursor]
    return {"week_key": week_key, "entries": entries}


@app.get("/api/admin/all")
async def admin_all(_: bool = Depends(require_admin)):
    cursor = entries_col.find({})
    entries = [entry_for_response(doc) async for doc in cursor]
    return {"entries": entries}


# Data-quality correction tool: reassigns an existing entry to a different
# roster identity (name + person_id only) without touching anything else
# about the activity. Exists for merging duplicate-person records created
# before roster-only enforcement existed — deliberately narrower than
# update_entry, which canonicalizes ward too and would overwrite historical
# ward text that must be preserved during a merge.
@app.patch("/api/admin/entries/{entry_id}/reassign-person", response_model=EntryOut)
async def admin_reassign_entry_person(
    entry_id: str, body: ReassignPersonIn, _: bool = Depends(require_admin)
):
    roster_person = await roster_col.find_one({"name_slug": slugify(body.name)})
    if not roster_person:
        raise HTTPException(400, "Target name must match an official roster entry")
    result = await entries_col.find_one_and_update(
        {"_id": ObjectId(entry_id)},
        {"$set": {"name": roster_person["name"], "person_id": roster_person["name_slug"]}},
        return_document=True,
    )
    if not result:
        raise HTTPException(404, "Entry not found")
    return entry_for_response(result)


@app.get("/api/admin/smartsheet/summary")
async def admin_smartsheet_summary(week_key: str, _: bool = Depends(require_admin)):
    cursor = entries_col.find({})
    entries = [entry_for_response(doc) async for doc in cursor]
    return summarize_smartsheet_entries(entries, week_key)


@app.get("/api/admin/smartsheet/review")
async def admin_smartsheet_review(_: bool = Depends(require_admin)):
    cursor = entries_col.find({})
    entries = [entry_for_response(doc) async for doc in cursor]
    return {"entries": review_entries(entries)}


@app.patch("/api/admin/entries/{entry_id}/smartsheet-category", response_model=EntryOut)
async def review_smartsheet_category(
    entry_id: str, body: CategoryReviewIn, _: bool = Depends(require_admin)
):
    category = body.smartsheet_category.strip().upper()
    if category not in REVIEWABLE_CATEGORIES:
        raise HTTPException(400, "Invalid SmartSheet category")
    result = await entries_col.find_one_and_update(
        {"_id": ObjectId(entry_id)},
        {"$set": {
            "smartsheet_category": category,
            "category_source": "admin_review",
            "category_reviewed": True,
            "category_reviewed_at": datetime.now(timezone.utc).isoformat(),
        }},
        return_document=True,
    )
    if not result:
        raise HTTPException(404, "Entry not found")
    return entry_for_response(result)


@app.get("/api/admin/smartsheet/export.csv")
async def admin_smartsheet_export_csv(
    week_key: str,
    category: str,
    _: bool = Depends(require_admin),
):
    normalized_category = category.strip().upper()
    allowed = set(REVIEWABLE_CATEGORIES) | {"ALL"}
    if normalized_category not in allowed:
        raise HTTPException(400, "Invalid SmartSheet export category")
    cursor = entries_col.find({"week_key": week_key})
    entries = [entry_for_response(doc) async for doc in cursor]
    csv_bytes = smartsheet_csv_bytes(entries, week_key, normalized_category, CONSTITUENCY)
    filename_category = {
        CANVASSING: "canvassing",
        PUBLIC_STREET_MEETING: "public-street-meetings",
        PRESENCE: "presence",
        "ALL": "all",
    }[normalized_category]
    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=smartsheet-{filename_category}-{week_key}.csv"},
    )


@app.get("/api/admin/smartsheet/export.xlsx")
async def admin_smartsheet_export_xlsx(
    week_key: str,
    category: str,
    _: bool = Depends(require_admin),
):
    normalized_category = category.strip().upper()
    cursor = entries_col.find({"week_key": week_key})
    entries = [entry_for_response(doc) async for doc in cursor]

    if normalized_category == "ALL":
        # "Download All Excel": one workbook, exactly the three category
        # worksheets — not the flat CSV "ALL" audit dump (that stays CSV-only).
        xlsx_bytes = smartsheet_workbook_all_categories_bytes(entries, week_key, CONSTITUENCY)
        filename = f"ntsikana-smartsheet-all-{week_key}.xlsx"
    elif normalized_category in REVIEWABLE_CATEGORIES:
        xlsx_bytes = smartsheet_xlsx_bytes(entries, week_key, normalized_category, CONSTITUENCY)
        filename_category = {
            CANVASSING: "canvassing",
            PUBLIC_STREET_MEETING: "public-street",
            PRESENCE: "presence",
        }[normalized_category]
        filename = f"ntsikana-smartsheet-{filename_category}-{week_key}.xlsx"
    else:
        raise HTTPException(400, "Invalid SmartSheet export category")

    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/admin/export.csv")
async def admin_export_csv(_: bool = Depends(require_admin)):
    cursor = entries_col.find({})
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["name", "ward", "week_label", "day", "activity_date", "type", "notes", "submitted_at"])
    async for doc in cursor:
        doc = enrich_entry(doc)
        writer.writerow([
            doc.get("name", ""), doc.get("ward", ""), doc.get("week_label", ""),
            doc.get("day", ""), doc.get("activity_date", ""), doc.get("type_display", doc.get("type", "")),
            doc.get("notes", "") or "", doc.get("submitted_at", ""),
        ])
    csv_bytes = buf.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=ward-tracker-export.csv"},
    )


@app.get("/api/admin/export.xlsx")
async def admin_export_xlsx(week_key: Optional[str] = None, _: bool = Depends(require_admin)):
    this_week_key = week_key or current_week_key()
    cursor = entries_col.find({"week_key": this_week_key})

    candidates = {}
    total_activities = 0
    type_counts = Counter()
    day_counts = {d: 0 for d in DAY_ORDER}
    async for doc in cursor:
        doc = enrich_entry(doc)
        total_activities += 1
        name = doc.get("name", "")
        day = doc.get("day", "")
        type_display = doc.get("type_display") or doc.get("type", "")
        notes = (doc.get("notes") or "").strip()

        if type_display:
            type_counts[type_display] += 1
        if day in day_counts:
            day_counts[day] += 1

        c = candidates.setdefault(name, {"ward": doc.get("ward", ""), "days": {}, "notes": {}})
        if not c["ward"]:
            c["ward"] = doc.get("ward", "")
        c["days"][day] = f'{c["days"][day]}, {type_display}' if day in c["days"] else type_display
        if notes:
            c["notes"][day] = f'{c["notes"][day]}; {notes}' if day in c["notes"] else notes

    total_candidates = len(candidates)
    names_sorted = sorted(candidates.keys(), key=lambda n: n.lower())
    roster_docs = [doc async for doc in roster_col.find({})]
    roster_size = len(roster_docs)
    breakdown_str = " | ".join(f"{t}: {n}" for t, n in sorted(type_counts.items()))

    day_dates = {d: activity_date_for_day_date(this_week_key, d) for d in DAY_ORDER}

    headers = (
        ["Name", "Ward"]
        + [f"{DAY_LABELS[d]}\n{day_dates[d].day} {MONTHS[day_dates[d].month - 1]}" for d in DAY_ORDER]
        + ["Notes"]
    )
    n_cols = len(headers)

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    if os.path.exists(LOGO_PATH):
        logo_img = XLImage(LOGO_PATH)
        logo_img.width = 50
        logo_img.height = 61
        ws.add_image(logo_img, "A1")
        ws.row_dimensions[1].height = 48

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
    title_cell = ws.cell(row=row, column=1, value="Ntsikana Constituency - Weekly Ward Activity Report")
    title_cell.font = Font(bold=True, size=14, color="153B63")
    row += 1

    summary_top_row = row
    ws.cell(row=row, column=1, value=f"Week: {format_week_label(this_week_key)}"); row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now(timezone.utc).strftime('%d %b %Y')}"); row += 1
    ws.cell(row=row, column=1, value=f"Total activities this week: {total_activities}"); row += 1
    ws.cell(row=row, column=1, value=f"Total candidates this week: {total_candidates}"); row += 1
    if breakdown_str:
        ws.cell(row=row, column=1, value=f"Activity breakdown: {breakdown_str}"); row += 1
    if roster_size > 0:
        ws.cell(row=row, column=1, value=f"Submission status: {total_candidates} of {roster_size} candidates submitted"); row += 1

    chart_col = n_cols + 2
    ws.cell(row=1, column=chart_col, value="Day")
    ws.cell(row=1, column=chart_col + 1, value="Count")
    for i, d in enumerate(DAY_ORDER):
        ws.cell(row=2 + i, column=chart_col, value=DAY_LABELS[d])
        ws.cell(row=2 + i, column=chart_col + 1, value=day_counts[d])
    ws.column_dimensions[get_column_letter(chart_col)].hidden = True
    ws.column_dimensions[get_column_letter(chart_col + 1)].hidden = True

    chart = BarChart()
    chart.type = "col"
    chart.title = "Activities per day"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Day"
    chart.legend = None
    chart.width = 8
    chart.height = 6
    chart_data = Reference(ws, min_col=chart_col + 1, min_row=1, max_row=1 + len(DAY_ORDER))
    chart_cats = Reference(ws, min_col=chart_col, min_row=2, max_row=1 + len(DAY_ORDER))
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(chart_cats)
    ws.add_chart(chart, f"{get_column_letter(chart_col)}{summary_top_row}")

    row += 1
    header_row_idx = row
    header_fill = PatternFill(start_color="2568AE", end_color="2568AE", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_side = Side(style="thin", color="DCD6C9")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    shade_fill = PatternFill(start_color="F3F1EC", end_color="F3F1EC", fill_type="solid")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_idx, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[header_row_idx].height = 30

    for row_offset, name in enumerate(names_sorted):
        r_idx = header_row_idx + 1 + row_offset
        c = candidates[name]
        row_fill = shade_fill if row_offset % 2 == 1 else None

        name_cell = ws.cell(row=r_idx, column=1, value=name)
        ward_cell = ws.cell(row=r_idx, column=2, value=c["ward"])
        for cell in (name_cell, ward_cell):
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill

        for day_idx, day in enumerate(DAY_ORDER):
            col = 3 + day_idx
            value = c["days"].get(day, "")
            cell = ws.cell(row=r_idx, column=col, value=value)
            cell.border = thin_border
            if row_fill:
                cell.fill = row_fill
            if value:
                cell.font = Font(bold=True)

        notes_parts = [f"{DAY_LABELS[d]}: {c['notes'][d]}" for d in DAY_ORDER if d in c["notes"]]
        notes_cell = ws.cell(row=r_idx, column=n_cols, value="; ".join(notes_parts))
        notes_cell.border = thin_border
        if row_fill:
            notes_cell.fill = row_fill

    last_row = header_row_idx + len(names_sorted)
    ws.freeze_panes = f"A{header_row_idx + 1}"

    for col_idx in range(1, n_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(part) for part in headers[col_idx - 1].split("\n"))
        for r_idx in range(header_row_idx + 1, last_row + 1):
            val = ws.cell(row=r_idx, column=col_idx).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = max_len + 2

    not_submitted = sorted(
        (r for r in roster_docs if not any(names_match(n, r.get("name", "")) for n in names_sorted)),
        key=lambda r: r.get("name", "").lower(),
    )
    if not_submitted:
        row = last_row + 2
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        section_title = ws.cell(row=row, column=1, value=f"Not Yet Submitted ({len(not_submitted)} of {roster_size})")
        section_title.font = Font(bold=True, size=12, color="B0473A")
        row += 1
        red_fill = PatternFill(start_color="FDF0EE", end_color="FDF0EE", fill_type="solid")
        for r in not_submitted:
            name_cell = ws.cell(row=row, column=1, value=r.get("name", ""))
            ward_cell = ws.cell(row=row, column=2, value=r.get("ward", ""))
            for cell in (name_cell, ward_cell):
                cell.fill = red_fill
                cell.border = thin_border
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ward-tracker-report.xlsx"},
    )


# ---------- Public: roster names (for name autocomplete) ----------
@app.get("/api/roster/names")
async def roster_names():
    cursor = roster_col.find({}, {"_id": 0, "name": 1, "ward": 1})
    return [doc async for doc in cursor]


# ---------- Admin: roster ----------
@app.get("/api/admin/roster")
async def get_roster(_: bool = Depends(require_admin)):
    cursor = roster_col.find({})
    return [oid_str(doc) async for doc in cursor]


@app.post("/api/admin/roster")
async def add_roster(body: RosterIn, _: bool = Depends(require_admin)):
    doc = body.model_dump()
    doc["name_slug"] = slugify(body.name)
    try:
        res = await roster_col.insert_one(doc)
    except Exception:
        raise HTTPException(409, "That candidate is already on the list")
    doc["_id"] = res.inserted_id
    return oid_str(doc)


@app.delete("/api/admin/roster/{roster_id}")
async def delete_roster(roster_id: str, _: bool = Depends(require_admin)):
    result = await roster_col.delete_one({"_id": ObjectId(roster_id)})
    if result.deleted_count == 0:
        raise HTTPException(404, "Not found")
    return {"deleted": True}


@app.get("/api/health")
async def health():
    return {"ok": True}
