import csv
import io
import re
from dataclasses import dataclass
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from week_dates import DAY_OFFSET, activity_date_for_day, format_week_label


CANVASSING = "CANVASSING"
PUBLIC_STREET_MEETING = "PUBLIC_STREET_MEETING"
PRESENCE = "PRESENCE"
NEEDS_REVIEW = "NEEDS_REVIEW"

CATEGORY_LABELS = {
    CANVASSING: "Canvassing Activities",
    PUBLIC_STREET_MEETING: "Public / Street Meetings",
    PRESENCE: "Presence Activities",
    NEEDS_REVIEW: "Needs Review",
}
REVIEWABLE_CATEGORIES = (CANVASSING, PUBLIC_STREET_MEETING, PRESENCE)

# Worksheet names for the XLSX SmartSheet workbook — short slugs, distinct from
# the longer admin-facing CATEGORY_LABELS used elsewhere in the UI.
SMARTSHEET_WORKSHEET_NAMES = {
    CANVASSING: "Canvassing",
    PUBLIC_STREET_MEETING: "Public-Street",
    PRESENCE: "Presence",
}

DEFAULT_CONSTITUENCY = "Ntsikana Constituency"

# Sentinel stored in an entry's `type` field when the candidate picked "Other"
# and typed a free-text activity. `type_display` then holds their exact wording;
# `type` stays this marker so classification never re-guesses that wording into
# a fixed category — it always resolves to NEEDS_REVIEW for an admin to decide.
CUSTOM_OTHER_TYPE = "Other"

SMARTSHEET_HEADERS = [
    "DATE",
    "TIME START",
    "TIME END",
    "CONSTITUENCY",
    "WARD",
    "VENUE",
    "ACTIVITY",
    "BOOST POST",
    "INFO GRAPHIC",
]
ALL_EXPORT_EXTRA_HEADERS = ["SMARTSHEET DESTINATION", "REVIEW STATUS"]

TIME_RE = re.compile(r"^(?:[01]\d|2[0-3]):[0-5]\d$")

CANONICAL_ACTIVITY_CATEGORY = {
    "Info table : canvassing": CANVASSING,
    "Canvassing Surgery": CANVASSING,
    "Door to Door": CANVASSING,
    "Telecanvassing": CANVASSING,
    "House Meeting": CANVASSING,
    "Info Table": CANVASSING,
    "Public Meeting": PUBLIC_STREET_MEETING,
    "Street Meeting": PUBLIC_STREET_MEETING,
    "Clean up": PRESENCE,
    "Oversight": PRESENCE,
    "Stakeholder meeting": PRESENCE,
    "Neighborhood watch patrol and handover": PRESENCE,
    "Rescue Event: Pothole repair": PRESENCE,
    "Rescue Event: Street Painting": PRESENCE,
    "Rescue Event: Lights": PRESENCE,
    "Hoot or Blue wave": PRESENCE,
    "Motorcade": PRESENCE,
    "Fun day": PRESENCE,
    "Sports day": PRESENCE,
    "Fundraiser": PRESENCE,
    "March": PRESENCE,
    "Picket": PRESENCE,
    "Rally": PRESENCE,
    "Soup Kitchen": PRESENCE,
    "Newspaper or Radio Advert": PRESENCE,
    "Religious Forum Address": PRESENCE,
    "Social Media": PRESENCE,
    "Poster removal": PRESENCE,
    "Women Safety": PRESENCE,
    "Fire extinguishers donated": PRESENCE,
    "Donation in kind": PRESENCE,
    "Blue Wave": PRESENCE,
    "Mayoral Campaign Pledges": PRESENCE,
    "Rescue Event": PRESENCE,
    "Care Event": PRESENCE,
    "Poster fighting": PRESENCE,
    "Leaflet Distribution": PRESENCE,
}


@dataclass(frozen=True)
class ActivityClassification:
    category: str
    canonical_activity: Optional[str]
    source: str
    needs_review: bool = False
    suggested_category: Optional[str] = None


def normalise_activity_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("&", " and ")
    text = re.sub(r"[-_/]+", " ", text)
    text = re.sub(r"[^a-z0-9:\s]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def entry_activity_text(doc: dict) -> str:
    return str(doc.get("type_display") or doc.get("type") or "").strip()


def is_custom_other_entry(doc: dict) -> bool:
    return str(doc.get("type") or "").strip() == CUSTOM_OTHER_TYPE


def normalize_time(value: object) -> Optional[str]:
    if value is None:
        return None
    value = str(value).strip()
    if value == "":
        return None
    if not TIME_RE.fullmatch(value):
        raise ValueError("Time must be in HH:MM 24-hour format")
    return value


def validate_time_range(start_time: object, end_time: object) -> tuple[Optional[str], Optional[str]]:
    start = normalize_time(start_time)
    end = normalize_time(end_time)
    if start and end and end <= start:
        raise ValueError("End Time must be after Start Time")
    return start, end


def normalise_venue(value: object) -> Optional[str]:
    if value is None:
        return None
    venue = str(value).strip()
    return venue or None


# Characters spreadsheet applications (Excel opening a CSV, or SmartSheet's own
# paste handler) may interpret as the start of a formula.
FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")


def spreadsheet_safe_text(value: object) -> str:
    """Neutralize spreadsheet formula injection in untrusted, candidate-typed
    export text (activity/venue/ward) without touching the stored database
    value — this only runs at the export/output boundary. A leading trigger
    character is prefixed with a single quote, the standard mitigation that
    keeps the value visibly readable as literal text instead of a formula."""
    text = str(value or "")
    if text and text[0] in FORMULA_TRIGGER_CHARS:
        return "'" + text
    return text


def activity_options() -> list[dict]:
    return [
        {"label": label, "category": category}
        for label, category in CANONICAL_ACTIVITY_CATEGORY.items()
    ]


_EXACT_ACTIVITY = {
    normalise_activity_text(label): (category, label)
    for label, category in CANONICAL_ACTIVITY_CATEGORY.items()
}

_ALIASES = {
    "door to door": "Door to Door",
    "door knocking": "Door to Door",
    "information table": "Info Table",
    "info table: canvassing": "Info table : canvassing",
    "info table canvassing": "Info table : canvassing",
    "soup kitchen": "Soup Kitchen",
    "blue wave": "Blue Wave",
    "hoot": "Hoot or Blue wave",
    "hooting": "Hoot or Blue wave",
    "hoot or bluewave": "Hoot or Blue wave",
    "clean up": "Clean up",
    "cleanup": "Clean up",
}

_KEYWORD_RULES = [
    (re.compile(r"\bpublic\s+meeting\b"), PUBLIC_STREET_MEETING, "Public Meeting"),
    (re.compile(r"\bstreet\s+meeting\b"), PUBLIC_STREET_MEETING, "Street Meeting"),
    (re.compile(r"\bhouse\s+meeting\b"), CANVASSING, "House Meeting"),
    (re.compile(r"\bdoor\s+to\s+door\b"), CANVASSING, "Door to Door"),
    (re.compile(r"\btele\s*canvass\w*\b"), CANVASSING, "Telecanvassing"),
    (re.compile(r"\binfo(?:rmation)?\s+table\b"), CANVASSING, "Info Table"),
    (re.compile(r"\bcanvass\w*\b"), CANVASSING, None),
    (re.compile(r"\bstakeholder\s+meeting\b"), PRESENCE, "Stakeholder meeting"),
    (re.compile(r"\bneighbou?rhood\s+watch\b"), PRESENCE, "Neighborhood watch patrol and handover"),
    (re.compile(r"\bpothole\s+repair\b|\brescue\s+event:\s*pothole\b"), PRESENCE, "Rescue Event: Pothole repair"),
    (re.compile(r"\bstreet\s+painting\b|\brescue\s+event:\s*street\s+painting\b"), PRESENCE, "Rescue Event: Street Painting"),
    (re.compile(r"\brescue\s+event:\s*lights\b"), PRESENCE, "Rescue Event: Lights"),
    (re.compile(r"\bblue\s+wave\b"), PRESENCE, "Blue Wave"),
    (re.compile(r"\bhoot(?:ing)?\b"), PRESENCE, "Hoot or Blue wave"),
    (re.compile(r"\bclean\s+up\b|\bcleanup\b|\bcleaning\s+up\b"), PRESENCE, "Clean up"),
    (re.compile(r"\boversight\b"), PRESENCE, "Oversight"),
    (re.compile(r"\bmotorcade\b"), PRESENCE, "Motorcade"),
    (re.compile(r"\bfun\s+day\b"), PRESENCE, "Fun day"),
    (re.compile(r"\bsports\s+day\b"), PRESENCE, "Sports day"),
    (re.compile(r"\bfundraiser\b"), PRESENCE, "Fundraiser"),
    (re.compile(r"\bmarch\b"), PRESENCE, "March"),
    (re.compile(r"\bpicket\b"), PRESENCE, "Picket"),
    (re.compile(r"\brally\b"), PRESENCE, "Rally"),
    (re.compile(r"\bsoup\s+kitchen\b"), PRESENCE, "Soup Kitchen"),
    (re.compile(r"\bnewspaper\b|\bradio\s+advert\b"), PRESENCE, "Newspaper or Radio Advert"),
    (re.compile(r"\breligious\s+forum\b"), PRESENCE, "Religious Forum Address"),
    (re.compile(r"\bsocial\s+media\b"), PRESENCE, "Social Media"),
    (re.compile(r"\bposter\s+removal\b"), PRESENCE, "Poster removal"),
    (re.compile(r"\bwomen\s+safety\b"), PRESENCE, "Women Safety"),
    (re.compile(r"\bfire\s+extinguishers?\s+donated\b"), PRESENCE, "Fire extinguishers donated"),
    (re.compile(r"\bdonation\s+in\s+kind\b"), PRESENCE, "Donation in kind"),
    (re.compile(r"\bmayoral\s+campaign\s+pledges?\b"), PRESENCE, "Mayoral Campaign Pledges"),
    (re.compile(r"\brescue\s+event\b"), PRESENCE, "Rescue Event"),
    (re.compile(r"\bcare\s+event\b"), PRESENCE, "Care Event"),
    (re.compile(r"\bposter\s+fighting\b"), PRESENCE, "Poster fighting"),
    (re.compile(r"\bpreparing\s+posters?\b|\bposter\s+preparation\b|\bposters?\b"), PRESENCE, None),
    (re.compile(r"\bleaflet\s+distribution\b"), PRESENCE, "Leaflet Distribution"),
]


def classify_activity_text(activity_text: object) -> ActivityClassification:
    text = normalise_activity_text(activity_text)
    if not text:
        return ActivityClassification(NEEDS_REVIEW, None, "automatic", True)

    exact = _EXACT_ACTIVITY.get(text)
    if exact:
        category, canonical = exact
        return ActivityClassification(category, canonical, "automatic")

    alias = _ALIASES.get(text)
    if alias:
        return ActivityClassification(
            CANONICAL_ACTIVITY_CATEGORY[alias], alias, "automatic"
        )

    for pattern, category, canonical in _KEYWORD_RULES:
        if pattern.search(text):
            return ActivityClassification(
                category=category,
                canonical_activity=canonical,
                source="automatic",
                needs_review=canonical is None,
                suggested_category=category if canonical is None else None,
            )

    return ActivityClassification(NEEDS_REVIEW, None, "automatic", True)


def classification_for_entry(doc: dict) -> ActivityClassification:
    stored_category = doc.get("smartsheet_category")
    stored_canonical = doc.get("canonical_activity")
    source = doc.get("category_source")

    if source == "admin_review" and stored_category in REVIEWABLE_CATEGORIES:
        canonical = stored_canonical if stored_canonical in CANONICAL_ACTIVITY_CATEGORY else None
        return ActivityClassification(stored_category, canonical, "admin_review")

    if stored_category in REVIEWABLE_CATEGORIES and stored_canonical in CANONICAL_ACTIVITY_CATEGORY:
        return ActivityClassification(stored_category, stored_canonical, source or "stored")

    if is_custom_other_entry(doc):
        return ActivityClassification(NEEDS_REVIEW, None, "automatic", True)

    return classify_activity_text(entry_activity_text(doc))


def reporting_metadata_for_submission(doc: dict, existing_doc: Optional[dict] = None) -> dict:
    activity_text = entry_activity_text(doc)
    if (
        existing_doc
        and existing_doc.get("category_source") == "admin_review"
        and normalise_activity_text(activity_text) == normalise_activity_text(entry_activity_text(existing_doc))
        and existing_doc.get("smartsheet_category") in REVIEWABLE_CATEGORIES
    ):
        return {
            "smartsheet_category": existing_doc.get("smartsheet_category"),
            "canonical_activity": existing_doc.get("canonical_activity"),
            "category_source": "admin_review",
            "category_reviewed": True,
            "category_reviewed_at": existing_doc.get("category_reviewed_at"),
        }

    if is_custom_other_entry(doc):
        return {
            "smartsheet_category": NEEDS_REVIEW,
            "canonical_activity": None,
            "category_source": "automatic",
            "category_reviewed": False,
            "category_reviewed_at": None,
        }

    classification = classify_activity_text(activity_text)
    return {
        "smartsheet_category": classification.category,
        "canonical_activity": classification.canonical_activity,
        "category_source": "automatic",
        "category_reviewed": False,
        "category_reviewed_at": None,
    }


def entry_activity_date(doc: dict) -> str:
    activity_date = doc.get("activity_date")
    if activity_date:
        return str(activity_date)
    week_key = doc.get("week_key")
    day = doc.get("day")
    if week_key and day in DAY_OFFSET:
        try:
            return activity_date_for_day(str(week_key), str(day))
        except ValueError:
            return ""
    return ""


def smartsheet_bucket(classification: ActivityClassification) -> str:
    if classification.needs_review or classification.category == NEEDS_REVIEW:
        return NEEDS_REVIEW
    return classification.category


def summarize_smartsheet_entries(entries: Iterable[dict], week_key: str) -> dict:
    all_entries = list(entries)
    overall = {category: 0 for category in (CANVASSING, PUBLIC_STREET_MEETING, PRESENCE, NEEDS_REVIEW)}
    weekly = {category: 0 for category in (CANVASSING, PUBLIC_STREET_MEETING, PRESENCE, NEEDS_REVIEW)}
    auto_categorized = 0
    manually_reviewed = 0

    for doc in all_entries:
        classification = classification_for_entry(doc)
        bucket = smartsheet_bucket(classification)
        overall[bucket] += 1
        if classification.source == "admin_review" and bucket != NEEDS_REVIEW:
            manually_reviewed += 1
        elif bucket != NEEDS_REVIEW:
            auto_categorized += 1
        if doc.get("week_key") == week_key:
            weekly[bucket] += 1

    return {
        "week_key": week_key,
        "week_label": format_week_label(week_key),
        "total_historical_activities": len(all_entries),
        "auto_categorized": auto_categorized,
        "manually_reviewed": manually_reviewed,
        "needs_review": overall[NEEDS_REVIEW],
        "weekly": {
            "total": sum(weekly.values()),
            CANVASSING: weekly[CANVASSING],
            PUBLIC_STREET_MEETING: weekly[PUBLIC_STREET_MEETING],
            PRESENCE: weekly[PRESENCE],
            NEEDS_REVIEW: weekly[NEEDS_REVIEW],
        },
    }


def review_payload_for_entry(doc: dict) -> Optional[dict]:
    classification = classification_for_entry(doc)
    if smartsheet_bucket(classification) != NEEDS_REVIEW:
        return None
    return {
        "id": str(doc.get("id") or doc.get("_id") or ""),
        "original_activity": entry_activity_text(doc),
        "name": doc.get("name", ""),
        "ward": doc.get("ward", ""),
        "activity_date": entry_activity_date(doc),
        "week_key": doc.get("week_key", ""),
        "week_label": format_week_label(doc["week_key"]) if doc.get("week_key") else "",
        "suggested_category": classification.suggested_category,
    }


def review_entries(entries: Iterable[dict]) -> list[dict]:
    rows = []
    for doc in entries:
        payload = review_payload_for_entry(doc)
        if payload:
            rows.append(payload)
    return sorted(rows, key=lambda row: (row.get("activity_date", ""), row.get("name", "")))


def _activity_for_export(doc: dict, classification: ActivityClassification, include_unreviewed: bool) -> str:
    if classification.canonical_activity:
        return classification.canonical_activity
    if classification.source == "admin_review" or include_unreviewed:
        return entry_activity_text(doc)
    return ""


def smartsheet_rows(
    entries: Iterable[dict],
    week_key: str,
    category: str,
    constituency: str = DEFAULT_CONSTITUENCY,
) -> list[list[str]]:
    category = category.upper()
    include_all = category == "ALL"
    if not include_all and category not in REVIEWABLE_CATEGORIES:
        raise ValueError("Invalid SmartSheet export category")

    rows = []
    for doc in entries:
        if doc.get("week_key") != week_key:
            continue

        classification = classification_for_entry(doc)
        bucket = smartsheet_bucket(classification)
        if not include_all and bucket != category:
            continue
        if not include_all and bucket == NEEDS_REVIEW:
            continue

        row = [
            entry_activity_date(doc),
            doc.get("start_time") or "",
            doc.get("end_time") or "",
            constituency,
            spreadsheet_safe_text(doc.get("ward") or ""),
            spreadsheet_safe_text(doc.get("venue") or ""),
            spreadsheet_safe_text(_activity_for_export(doc, classification, include_all)),
            "",
            "",
        ]
        if include_all:
            destination = CATEGORY_LABELS.get(bucket, CATEGORY_LABELS[NEEDS_REVIEW])
            status = "Needs review" if bucket == NEEDS_REVIEW else "Ready"
            row.extend([destination, status])
        rows.append(row)

    return sorted(rows, key=lambda row: (row[0], row[1], row[4], row[6]))


def smartsheet_csv_bytes(
    entries: Iterable[dict],
    week_key: str,
    category: str,
    constituency: str = DEFAULT_CONSTITUENCY,
) -> bytes:
    category = category.upper()
    headers = list(SMARTSHEET_HEADERS)
    if category == "ALL":
        headers.extend(ALL_EXPORT_EXTRA_HEADERS)

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(smartsheet_rows(entries, week_key, category, constituency))
    return buf.getvalue().encode("utf-8-sig")


def _write_smartsheet_worksheet(ws, rows: list[list[str]]) -> None:
    """Plain, paste-friendly worksheet: header row 1, data from row 2, no
    title/metadata rows, no merged cells, no formulas — one value per cell."""
    ws.append(SMARTSHEET_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"

    for col_idx, header in enumerate(SMARTSHEET_HEADERS, start=1):
        max_len = len(header)
        for row in rows:
            value = row[col_idx - 1]
            if value:
                max_len = max(max_len, len(str(value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def smartsheet_xlsx_bytes(
    entries: Iterable[dict],
    week_key: str,
    category: str,
    constituency: str = DEFAULT_CONSTITUENCY,
) -> bytes:
    """A single-worksheet .xlsx for one SmartSheet category — same
    classification/rows as the CSV export, just genuine Excel cells."""
    category = category.upper()
    if category not in REVIEWABLE_CATEGORIES:
        raise ValueError("Invalid SmartSheet export category")

    rows = smartsheet_rows(entries, week_key, category, constituency)
    wb = Workbook()
    ws = wb.active
    ws.title = SMARTSHEET_WORKSHEET_NAMES[category]
    _write_smartsheet_worksheet(ws, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def smartsheet_workbook_all_categories_bytes(
    entries: Iterable[dict],
    week_key: str,
    constituency: str = DEFAULT_CONSTITUENCY,
) -> bytes:
    """"Download All Excel": one workbook, exactly three worksheets
    (Canvassing, Public-Street, Presence), each identical in shape to the
    single-category export. NEEDS_REVIEW entries never appear here."""
    wb = Workbook()
    wb.remove(wb.active)  # drop openpyxl's default blank sheet
    for category in (CANVASSING, PUBLIC_STREET_MEETING, PRESENCE):
        rows = smartsheet_rows(entries, week_key, category, constituency)
        ws = wb.create_sheet(SMARTSHEET_WORKSHEET_NAMES[category])
        _write_smartsheet_worksheet(ws, rows)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
