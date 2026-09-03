import asyncio
import copy
import csv
import io
import os
import unittest
from types import SimpleNamespace


try:
    import fastapi  # noqa: F401
    from bson import ObjectId
    from fastapi import HTTPException

    HAS_API_DEPS = True
except ModuleNotFoundError:
    ObjectId = None
    HTTPException = Exception
    HAS_API_DEPS = False

try:
    from fastapi.testclient import TestClient  # requires httpx

    HAS_TESTCLIENT = True
except ImportError:
    TestClient = None
    HAS_TESTCLIENT = False


@unittest.skipUnless(HAS_API_DEPS, "FastAPI dependencies are not installed")
class FastApiSmartSheetTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("MONGO_URI", "mongodb://127.0.0.1:1")
        os.environ.setdefault("ADMIN_PIN", "1234")
        os.environ.setdefault("JWT_SECRET", "local-test-secret")

        global appmod
        import main as appmod

    def setUp(self):
        self.original_entries_col = appmod.entries_col
        self.original_roster_col = appmod.roster_col
        self.entries = FakeCollection()
        self.roster = FakeCollection()
        self.roster.docs = [
            {"_id": ObjectId(), "name": "Test Candidate", "ward": "Ward 1", "name_slug": "test-candidate"},
            {"_id": ObjectId(), "name": "Second Candidate", "ward": "Ward 2", "name_slug": "second-candidate"},
        ]
        appmod.entries_col = self.entries
        appmod.roster_col = self.roster

    def tearDown(self):
        appmod.entries_col = self.original_entries_col
        appmod.roster_col = self.original_roster_col

    def test_candidate_submission_persists_new_fields_and_derived_category(self):
        body = appmod.EntryIn(
            person_id="test-candidate",
            name="Test Candidate",
            ward="Ward 1",
            day="mon",
            type="Door to Door",
            type_display="Door to Door",
            notes=None,
            week_key="2026-08-30",
            week_label="31 Aug - 6 Sep",
            activity_date="2026-08-31",
            start_time="09:00",
            end_time="10:30",
            venue="Ward office",
        )

        result = asyncio.run(appmod.create_entry(body))

        self.assertEqual(len(self.entries.docs), 1)
        self.assertEqual(result["start_time"], "09:00")
        self.assertEqual(result["end_time"], "10:30")
        self.assertEqual(result["venue"], "Ward office")
        self.assertEqual(result["smartsheet_category"], "CANVASSING")
        self.assertEqual(result["canonical_activity"], "Door to Door")

    def test_candidate_submission_rejects_obvious_invalid_time_range(self):
        body = appmod.EntryIn(
            person_id="test-candidate",
            name="Test Candidate",
            ward="Ward 1",
            day="mon",
            type="Blue Wave",
            type_display="Blue Wave",
            week_key="2026-08-30",
            week_label="31 Aug - 6 Sep",
            activity_date="2026-08-31",
            start_time="11:00",
            end_time="10:30",
            venue="Main road",
        )

        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.create_entry(body))
        self.assertEqual(exc.exception.status_code, 400)
        self.assertEqual(len(self.entries.docs), 0)

    def test_update_without_notes_preserves_existing_notes(self):
        entry_id = ObjectId()
        self.entries.docs = [
            entry_doc(
                _id=entry_id,
                type_display="Door to Door",
                week_key="2026-08-30",
                day="mon",
                start_time="09:00",
                end_time="10:00",
                venue="Area 1",
            )
        ]
        self.entries.docs[0]["notes"] = "Existing historical note"
        body = appmod.EntryIn(
            person_id="test-candidate",
            name="Test Candidate",
            ward="Ward 1",
            day="mon",
            type="Door to Door",
            type_display="Door to Door",
            notes=None,
            week_key="2026-08-30",
            week_label="31 Aug - 6 Sep",
            activity_date="2026-08-31",
            start_time="09:30",
            end_time="10:30",
            venue="Area 1",
        )

        updated = asyncio.run(appmod.update_entry(str(entry_id), body))

        self.assertEqual(updated["notes"], "Existing historical note")
        self.assertEqual(self.entries.docs[0]["notes"], "Existing historical note")

    def test_existing_csv_and_xlsx_exports_run_against_mocked_data(self):
        self.entries.docs = [
            entry_doc(type_display="Door to Door", week_key="2026-08-30", day="mon"),
        ]
        self.roster.docs = [
            {"_id": ObjectId(), "name": "Test Candidate", "ward": "Ward 1", "name_slug": "test-candidate"}
        ]

        csv_response = asyncio.run(appmod.admin_export_csv(True))
        csv_payload = asyncio.run(streaming_body(csv_response)).decode("utf-8-sig")
        self.assertIn("name,ward,week_label,day,activity_date,type,notes,submitted_at", csv_payload)
        self.assertIn("Door to Door", csv_payload)

        xlsx_response = asyncio.run(appmod.admin_export_xlsx("2026-08-30", True))
        xlsx_payload = asyncio.run(streaming_body(xlsx_response))
        self.assertTrue(xlsx_payload.startswith(b"PK"))

    def test_smartsheet_summary_export_and_review_use_mocked_data_only(self):
        unclear_id = ObjectId()
        self.entries.docs = [
            entry_doc(type_display="Door to Door", week_key="2026-08-30", day="mon", start_time="09:00", end_time="10:00", venue="Area 1"),
            entry_doc(type_display="Street Meeting", week_key="2026-08-30", day="tue"),
            entry_doc(type_display="Blue Wave", week_key="2026-08-30", day="wed"),
            entry_doc(_id=unclear_id, type_display="Meeting", type="Meeting", week_key="2026-08-30", day="thu"),
        ]

        summary = asyncio.run(appmod.admin_smartsheet_summary("2026-08-30", True))
        self.assertEqual(summary["weekly"]["CANVASSING"], 1)
        self.assertEqual(summary["weekly"]["PUBLIC_STREET_MEETING"], 1)
        self.assertEqual(summary["weekly"]["PRESENCE"], 1)
        self.assertEqual(summary["weekly"]["NEEDS_REVIEW"], 1)

        canvassing_response = asyncio.run(appmod.admin_smartsheet_export_csv("2026-08-30", "CANVASSING", True))
        canvassing_rows = csv_rows(asyncio.run(streaming_body(canvassing_response)))
        self.assertEqual(canvassing_rows[0][:9], [
            "DATE", "TIME START", "TIME END", "CONSTITUENCY", "WARD", "VENUE", "ACTIVITY", "BOOST POST", "INFO GRAPHIC",
        ])
        self.assertEqual(len(canvassing_rows), 2)
        self.assertEqual(canvassing_rows[1][6], "Door to Door")

        review_rows = asyncio.run(appmod.admin_smartsheet_review(True))["entries"]
        self.assertEqual(len(review_rows), 1)
        self.assertEqual(review_rows[0]["original_activity"], "Meeting")

        reviewed = asyncio.run(appmod.review_smartsheet_category(
            str(unclear_id),
            appmod.CategoryReviewIn(smartsheet_category="PRESENCE"),
            True,
        ))
        self.assertEqual(reviewed["type_display"], "Meeting")
        self.assertEqual(reviewed["smartsheet_category"], "PRESENCE")
        self.assertEqual(self.entries.docs[-1]["type_display"], "Meeting")

    def test_smartsheet_xlsx_export_uses_same_classification_as_csv(self):
        self.entries.docs = [
            entry_doc(type_display="Door to Door", week_key="2026-08-30", day="mon", start_time="09:00", end_time="10:00", venue="Area 1"),
            entry_doc(type_display="Street Meeting", week_key="2026-08-30", day="tue"),
            entry_doc(type_display="Blue Wave", week_key="2026-08-30", day="wed"),
            entry_doc(type_display="Meeting", type="Meeting", week_key="2026-08-30", day="thu"),
        ]

        xlsx_response = asyncio.run(appmod.admin_smartsheet_export_xlsx("2026-08-30", "CANVASSING", True))
        xlsx_payload = asyncio.run(streaming_body(xlsx_response))
        self.assertTrue(xlsx_payload.startswith(b"PK"), "must be a genuine .xlsx, not CSV bytes")

        csv_response = asyncio.run(appmod.admin_smartsheet_export_csv("2026-08-30", "CANVASSING", True))
        csv_rows_ = csv_rows(asyncio.run(streaming_body(csv_response)))

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx_payload))
        self.assertEqual(wb.sheetnames, ["Canvassing"])
        ws = wb.active
        xlsx_rows = [[c.value or "" for c in row] for row in ws.iter_rows()]
        # Same one Door to Door row, same 9-column shape, in both formats.
        self.assertEqual(len(xlsx_rows) - 1, len(csv_rows_) - 1)
        self.assertEqual(xlsx_rows[1][6], csv_rows_[1][6])

        all_xlsx = asyncio.run(streaming_body(asyncio.run(appmod.admin_smartsheet_export_xlsx("2026-08-30", "ALL", True))))
        wb_all = load_workbook(io.BytesIO(all_xlsx))
        self.assertEqual(wb_all.sheetnames, ["Canvassing", "Public-Street", "Presence"])

    def test_admin_auth_rejects_missing_token_and_accepts_generated_token(self):
        with self.assertRaises(HTTPException):
            asyncio.run(appmod.require_admin(None))

        token = appmod.make_admin_token()
        self.assertTrue(asyncio.run(appmod.require_admin("Bearer " + token)))

    def _other_entry_body(self, type_display, **overrides):
        kwargs = dict(
            person_id="test-candidate",
            name="Test Candidate",
            ward="Ward 1",
            day="mon",
            type="Other",
            type_display=type_display,
            notes=None,
            week_key="2026-08-30",
            week_label="31 Aug - 6 Sep",
            activity_date="2026-08-31",
            start_time="09:00",
            end_time="10:30",
            venue="Ward office",
        )
        kwargs.update(overrides)
        return appmod.EntryIn(**kwargs)

    def test_other_submission_preserves_custom_wording_and_needs_review(self):
        result = asyncio.run(appmod.create_entry(self._other_entry_body("Community prayer event")))

        self.assertEqual(result["type"], "Other")
        self.assertEqual(result["type_display"], "Community prayer event")
        self.assertEqual(result["smartsheet_category"], "NEEDS_REVIEW")
        self.assertIsNone(result["canonical_activity"])

    def test_other_submission_matching_official_label_is_not_auto_classified(self):
        # Candidate deliberately typed Other with wording that matches a real
        # activity — must stay NEEDS_REVIEW, not silently become CANVASSING.
        result = asyncio.run(appmod.create_entry(self._other_entry_body("Door to Door")))

        self.assertEqual(result["type"], "Other")
        self.assertEqual(result["type_display"], "Door to Door")
        self.assertEqual(result["smartsheet_category"], "NEEDS_REVIEW")
        self.assertIsNone(result["canonical_activity"])

    def test_other_submission_rejects_blank_text(self):
        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.create_entry(self._other_entry_body("")))
        self.assertEqual(exc.exception.status_code, 400)
        self.assertEqual(len(self.entries.docs), 0)

    def test_other_submission_rejects_whitespace_only_text(self):
        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.create_entry(self._other_entry_body("   \t  ")))
        self.assertEqual(exc.exception.status_code, 400)
        self.assertEqual(len(self.entries.docs), 0)

    def test_other_submission_trims_leading_and_trailing_whitespace(self):
        result = asyncio.run(appmod.create_entry(self._other_entry_body("  Community prayer event  ")))
        self.assertEqual(result["type_display"], "Community prayer event")

    def test_other_entry_appears_in_admin_review_workflow(self):
        asyncio.run(appmod.create_entry(self._other_entry_body("Community prayer event")))

        review_rows = asyncio.run(appmod.admin_smartsheet_review(True))["entries"]
        self.assertEqual(len(review_rows), 1)
        self.assertEqual(review_rows[0]["original_activity"], "Community prayer event")

        summary = asyncio.run(appmod.admin_smartsheet_summary("2026-08-30", True))
        self.assertEqual(summary["weekly"]["NEEDS_REVIEW"], 1)
        self.assertEqual(summary["weekly"]["CANVASSING"], 0)

    def test_other_entry_excluded_from_category_export_until_reviewed(self):
        # "Door to Door" wording via Other must not leak into the Canvassing export.
        asyncio.run(appmod.create_entry(self._other_entry_body("Door to Door")))
        asyncio.run(appmod.create_entry(appmod.EntryIn(
            person_id="test-candidate-2", name="Second Candidate", ward="Ward 2", day="tue",
            type="Door to Door", type_display="Door to Door",
            week_key="2026-08-30", week_label="31 Aug - 6 Sep", activity_date="2026-09-01",
            start_time="09:00", end_time="10:00", venue="Ward office",
        )))

        canvassing_response = asyncio.run(appmod.admin_smartsheet_export_csv("2026-08-30", "CANVASSING", True))
        canvassing_rows = csv_rows(asyncio.run(streaming_body(canvassing_response)))
        # Header + exactly the one genuine (non-Other) Door to Door submission.
        self.assertEqual(len(canvassing_rows), 2)

        all_response = asyncio.run(appmod.admin_smartsheet_export_csv("2026-08-30", "ALL", True))
        all_rows = csv_rows(asyncio.run(streaming_body(all_response)))
        needs_review = [r for r in all_rows[1:] if r[-2:] == ["Needs Review", "Needs review"]]
        self.assertEqual(len(needs_review), 1)
        self.assertEqual(needs_review[0][6], "Door to Door")

    def test_editing_existing_other_entry_preserves_notes_time_venue(self):
        entry_id = ObjectId()
        self.entries.docs = [
            entry_doc(
                _id=entry_id,
                type="Other",
                type_display="Community prayer event",
                week_key="2026-08-30",
                day="mon",
                start_time="09:00",
                end_time="10:00",
                venue="Church hall",
            )
        ]
        self.entries.docs[0]["notes"] = "Well attended"
        self.entries.docs[0]["smartsheet_category"] = "NEEDS_REVIEW"
        self.entries.docs[0]["canonical_activity"] = None
        self.entries.docs[0]["category_source"] = "automatic"

        body = self._other_entry_body(
            "Community prayer event",
            notes=None,
            start_time="09:30",
            end_time="10:30",
            venue="Church hall",
        )
        updated = asyncio.run(appmod.update_entry(str(entry_id), body))

        self.assertEqual(updated["notes"], "Well attended")
        self.assertEqual(updated["start_time"], "09:30")
        self.assertEqual(updated["end_time"], "10:30")
        self.assertEqual(updated["venue"], "Church hall")
        self.assertEqual(updated["type_display"], "Community prayer event")
        self.assertEqual(updated["smartsheet_category"], "NEEDS_REVIEW")

    def test_editing_admin_reviewed_other_entry_keeps_override_when_unchanged(self):
        entry_id = ObjectId()
        self.entries.docs = [
            entry_doc(
                _id=entry_id,
                type="Other",
                type_display="Community prayer event",
                week_key="2026-08-30",
                day="mon",
                start_time="09:00",
                end_time="10:00",
                venue="Church hall",
            )
        ]
        self.entries.docs[0]["smartsheet_category"] = "PRESENCE"
        self.entries.docs[0]["category_source"] = "admin_review"
        self.entries.docs[0]["category_reviewed_at"] = "2026-09-01T10:00:00+00:00"

        body = self._other_entry_body(
            "Community prayer event", notes=None, venue="Church hall",
        )
        updated = asyncio.run(appmod.update_entry(str(entry_id), body))

        self.assertEqual(updated["smartsheet_category"], "PRESENCE")
        self.assertEqual(updated["category_source"], "admin_review")


@unittest.skipUnless(HAS_API_DEPS, "FastAPI dependencies are not installed")
class RosterOnlyIdentityTests(unittest.TestCase):
    """Follow-up: candidates must select a real roster person; the backend
    must enforce this independent of the UI, and canonicalize name/ward/
    person_id from the roster so spelling/case/abbreviation variants of the
    same person can never create a second dashboard identity (the bug behind
    "Cecilia Anne Auld (CLLR)" vs "Cecilia Auld" showing as two people)."""

    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("MONGO_URI", "mongodb://127.0.0.1:1")
        os.environ.setdefault("ADMIN_PIN", "1234")
        os.environ.setdefault("JWT_SECRET", "local-test-secret")
        global appmod
        import main as appmod

    def setUp(self):
        self.original_entries_col = appmod.entries_col
        self.original_roster_col = appmod.roster_col
        self.entries = FakeCollection()
        self.roster = FakeCollection()
        self.roster.docs = [
            {"_id": ObjectId(), "name": "Cecilia Anne Auld (CLLR)", "ward": "Ward 4", "name_slug": "cecilia-anne-auld-cllr"},
        ]
        appmod.entries_col = self.entries
        appmod.roster_col = self.roster

    def tearDown(self):
        appmod.entries_col = self.original_entries_col
        appmod.roster_col = self.original_roster_col

    def _body(self, **overrides):
        kwargs = dict(
            person_id="whatever-the-client-sends",
            name="Cecilia Anne Auld (CLLR)",
            ward="Wrong Ward",
            day="mon",
            type="Door to Door",
            type_display="Door to Door",
            week_key="2026-08-30",
            week_label="31 Aug - 6 Sep",
            activity_date="2026-08-31",
            start_time="09:00",
            end_time="10:00",
            venue="Ward office",
        )
        kwargs.update(overrides)
        return appmod.EntryIn(**kwargs)

    def test_create_entry_rejects_name_not_on_roster(self):
        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.create_entry(self._body(name="Someone Not On The Roster")))
        self.assertEqual(exc.exception.status_code, 400)
        self.assertEqual(len(self.entries.docs), 0)

    def test_create_entry_ignores_client_supplied_ward_and_person_id(self):
        result = asyncio.run(appmod.create_entry(self._body()))
        self.assertEqual(self.entries.docs[0]["ward"], "Ward 4")
        self.assertEqual(self.entries.docs[0]["person_id"], "cecilia-anne-auld-cllr")
        self.assertEqual(self.entries.docs[0]["name"], "Cecilia Anne Auld (CLLR)")

    def test_create_entry_canonicalizes_case_and_whitespace_variants(self):
        result = asyncio.run(appmod.create_entry(self._body(name="  cecilia anne auld (cllr)  ")))
        self.assertEqual(self.entries.docs[0]["name"], "Cecilia Anne Auld (CLLR)")
        self.assertEqual(self.entries.docs[0]["person_id"], "cecilia-anne-auld-cllr")

    def test_spelling_variant_cannot_create_a_second_person(self):
        # "Cecilia Auld" is not the exact roster entry -> must be rejected,
        # not silently accepted as a new/different identity.
        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.create_entry(self._body(name="Cecilia Auld")))
        self.assertEqual(exc.exception.status_code, 400)
        self.assertEqual(len(self.entries.docs), 0)

    def test_two_submissions_of_the_same_roster_person_share_one_person_id(self):
        asyncio.run(appmod.create_entry(self._body(day="mon")))
        asyncio.run(appmod.create_entry(self._body(name="CECILIA ANNE AULD (CLLR)", day="tue", activity_date=None)))
        person_ids = {doc["person_id"] for doc in self.entries.docs}
        self.assertEqual(person_ids, {"cecilia-anne-auld-cllr"})

    def test_update_entry_also_enforces_roster(self):
        entry_id = ObjectId()
        self.entries.docs = [
            entry_doc(_id=entry_id, type_display="Door to Door", week_key="2026-08-30", day="mon")
        ]
        self.entries.docs[0]["person_id"] = "cecilia-anne-auld-cllr"
        self.entries.docs[0]["name"] = "Cecilia Anne Auld (CLLR)"

        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.update_entry(str(entry_id), self._body(
                person_id="cecilia-anne-auld-cllr", name="Someone Else Entirely",
            )))
        self.assertEqual(exc.exception.status_code, 400)
        self.assertEqual(self.entries.docs[0]["name"], "Cecilia Anne Auld (CLLR)")

    def test_update_entry_with_valid_roster_name_still_works(self):
        entry_id = ObjectId()
        self.entries.docs = [
            entry_doc(_id=entry_id, type_display="Door to Door", week_key="2026-08-30", day="mon")
        ]
        self.entries.docs[0]["person_id"] = "cecilia-anne-auld-cllr"
        self.entries.docs[0]["name"] = "Cecilia Anne Auld (CLLR)"

        updated = asyncio.run(appmod.update_entry(str(entry_id), self._body(
            person_id="cecilia-anne-auld-cllr", start_time="10:00", end_time="11:00",
        )))
        self.assertEqual(updated["start_time"], "10:00")

    def test_reassign_person_moves_identity_without_touching_anything_else(self):
        entry_id = ObjectId()
        self.entries.docs = [entry_doc(_id=entry_id, type_display="Door to Door", week_key="2026-08-23", day="mon")]
        original = dict(self.entries.docs[0])
        original["person_id"] = "cecilia-auld"
        original["name"] = "Cecilia Auld"
        original["ward"] = "Ward 13 and Ward 7 RMM"
        original["notes"] = "Original historical note"
        original["smartsheet_category"] = "CANVASSING"
        original["category_source"] = "automatic"
        self.entries.docs[0] = original

        result = asyncio.run(appmod.admin_reassign_entry_person(
            str(entry_id), appmod.ReassignPersonIn(name="Cecilia Anne Auld (CLLR)"), True,
        ))

        self.assertEqual(result["name"], "Cecilia Anne Auld (CLLR)")
        self.assertEqual(result["person_id"], "cecilia-anne-auld-cllr")
        # Everything else must be byte-for-byte unchanged, most importantly ward.
        self.assertEqual(result["ward"], "Ward 13 and Ward 7 RMM")
        self.assertEqual(result["notes"], "Original historical note")
        self.assertEqual(result["smartsheet_category"], "CANVASSING")
        self.assertEqual(result["day"], "mon")
        self.assertEqual(result["week_key"], "2026-08-23")
        self.assertEqual(result["type_display"], "Door to Door")

    def test_reassign_person_rejects_target_not_on_roster(self):
        entry_id = ObjectId()
        self.entries.docs = [entry_doc(_id=entry_id, type_display="Door to Door", week_key="2026-08-30", day="mon")]
        self.entries.docs[0]["name"] = "Cecilia Auld"
        self.entries.docs[0]["person_id"] = "cecilia-auld"

        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.admin_reassign_entry_person(
                str(entry_id), appmod.ReassignPersonIn(name="Not On The Roster"), True,
            ))
        self.assertEqual(exc.exception.status_code, 400)
        self.assertEqual(self.entries.docs[0]["name"], "Cecilia Auld")

    def test_reassign_person_requires_admin(self):
        with self.assertRaises(HTTPException):
            asyncio.run(appmod.require_admin(None))


@unittest.skipUnless(HAS_API_DEPS and HAS_TESTCLIENT, "FastAPI TestClient (httpx) is not installed")
class RosterOnlyIdentityApiTests(unittest.TestCase):
    """Real HTTP round-trip: a direct API request that never touched the UI
    must still be rejected if it names a person not on the roster."""

    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("MONGO_URI", "mongodb://127.0.0.1:1")
        os.environ.setdefault("ADMIN_PIN", "1234")
        os.environ.setdefault("JWT_SECRET", "local-test-secret")
        global appmod
        import main as appmod

    def setUp(self):
        self.original_entries_col = appmod.entries_col
        self.original_roster_col = appmod.roster_col
        appmod.entries_col = FakeCollection()
        appmod.roster_col = FakeCollection()
        appmod.roster_col.docs = [
            {"_id": ObjectId(), "name": "Marnich Roets", "ward": "Ward 1", "name_slug": "marnich-roets"},
        ]
        self.client = TestClient(appmod.app)
        self.client.__enter__()

    def tearDown(self):
        self.client.__exit__(None, None, None)
        appmod.entries_col = self.original_entries_col
        appmod.roster_col = self.original_roster_col

    def test_direct_api_post_with_unregistered_person_is_rejected(self):
        response = self.client.post("/api/entries", json=dict(
            person_id="attacker-chosen-id", name="Totally Made Up Person", ward="Nowhere",
            day="mon", type="Door to Door", type_display="Door to Door", notes=None,
            week_key="2026-08-30", week_label="31 Aug - 6 Sep", activity_date="2026-08-31",
            start_time="09:00", end_time="10:00", venue="Ward office",
        ))
        self.assertEqual(response.status_code, 400)

    def test_direct_api_post_with_registered_roster_name_succeeds_and_canonicalizes(self):
        response = self.client.post("/api/entries", json=dict(
            person_id="attacker-chosen-id", name="marnich roets", ward="Somewhere Else",
            day="mon", type="Door to Door", type_display="Door to Door", notes=None,
            week_key="2026-08-30", week_label="31 Aug - 6 Sep", activity_date="2026-08-31",
            start_time="09:00", end_time="10:00", venue="Ward office",
        ))
        self.assertEqual(response.status_code, 200)
        stored = appmod.entries_col.docs[0]
        self.assertEqual(stored["person_id"], "marnich-roets")
        self.assertEqual(stored["name"], "Marnich Roets")
        self.assertEqual(stored["ward"], "Ward 1")

    def test_reassign_person_endpoint_requires_admin_token(self):
        created = self.client.post("/api/entries", json=dict(
            person_id="x", name="marnich roets", ward="Ward 1", day="mon",
            type="Door to Door", type_display="Door to Door", notes=None,
            week_key="2026-08-30", week_label="31 Aug - 6 Sep", activity_date="2026-08-31",
            start_time="09:00", end_time="10:00", venue="Ward office",
        )).json()
        response = self.client.patch(
            f"/api/admin/entries/{created['id']}/reassign-person",
            json={"name": "Marnich Roets"},
        )
        self.assertEqual(response.status_code, 401)

    def test_reassign_person_endpoint_with_admin_token_preserves_ward(self):
        appmod.roster_col.docs.append(
            {"_id": ObjectId(), "name": "Cecilia Anne Auld (CLLR)", "ward": "Raymond Mhlaba", "name_slug": "cecilia-anne-auld-cllr"}
        )
        created = self.client.post("/api/entries", json=dict(
            person_id="x", name="marnich roets", ward="Ward 1", day="mon",
            type="Door to Door", type_display="Door to Door", notes=None,
            week_key="2026-08-30", week_label="31 Aug - 6 Sep", activity_date="2026-08-31",
            start_time="09:00", end_time="10:00", venue="Ward office",
        )).json()
        token = appmod.make_admin_token()
        response = self.client.patch(
            f"/api/admin/entries/{created['id']}/reassign-person",
            json={"name": "Cecilia Anne Auld (CLLR)"},
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["name"], "Cecilia Anne Auld (CLLR)")
        self.assertEqual(body["person_id"], "cecilia-anne-auld-cllr")
        self.assertEqual(body["ward"], "Ward 1", "ward must stay whatever it was on the original entry")


@unittest.skipUnless(HAS_API_DEPS and HAS_TESTCLIENT, "FastAPI TestClient (httpx) is not installed")
class CandidateApiExposureTests(unittest.TestCase):
    """Follow-up audit: candidate-facing entry endpoints must never return
    SmartSheet/admin-only fields over the wire, regardless of what the UI
    renders. Exercised through the real ASGI request/response cycle (not a
    direct function call) so response_model filtering is actually verified."""

    INTERNAL_FIELDS = {
        "smartsheet_category", "canonical_activity", "category_source",
        "category_reviewed", "category_reviewed_at",
    }
    CANDIDATE_FIELDS = {
        "id", "day", "type", "type_display", "notes",
        "week_key", "activity_date", "start_time", "end_time", "venue",
    }

    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("MONGO_URI", "mongodb://127.0.0.1:1")
        os.environ.setdefault("ADMIN_PIN", "1234")
        os.environ.setdefault("JWT_SECRET", "local-test-secret")
        global appmod
        import main as appmod

    def setUp(self):
        self.original_entries_col = appmod.entries_col
        self.original_roster_col = appmod.roster_col
        appmod.entries_col = FakeCollection()
        appmod.roster_col = FakeCollection()
        appmod.roster_col.docs = [
            {"_id": ObjectId(), "name": "Test Candidate", "ward": "Ward 1", "name_slug": "test-candidate"},
        ]
        self.client = TestClient(appmod.app)
        self.client.__enter__()

    def tearDown(self):
        self.client.__exit__(None, None, None)
        appmod.entries_col = self.original_entries_col
        appmod.roster_col = self.original_roster_col

    def _submit(self, **overrides):
        body = dict(
            person_id="test-candidate", name="Test Candidate", ward="Ward 1", day="mon",
            type="Door to Door", type_display="Door to Door", notes=None,
            week_key="2026-08-30", week_label="31 Aug - 6 Sep", activity_date="2026-08-31",
            start_time="09:00", end_time="10:30", venue="Ward office",
        )
        body.update(overrides)
        return self.client.post("/api/entries", json=body)

    def test_post_response_excludes_internal_smartsheet_fields(self):
        response = self._submit()
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertFalse(self.INTERNAL_FIELDS & body.keys(), body.keys())
        self.assertEqual(set(body.keys()), self.CANDIDATE_FIELDS)

    def test_get_response_excludes_internal_smartsheet_fields(self):
        self._submit()
        response = self.client.get("/api/entries", params={"person_id": "test-candidate", "week_key": "2026-08-30"})
        self.assertEqual(response.status_code, 200)
        entries = response.json()
        self.assertEqual(len(entries), 1)
        self.assertFalse(self.INTERNAL_FIELDS & entries[0].keys(), entries[0].keys())

    def test_get_response_still_contains_every_field_the_frontend_needs(self):
        self._submit()
        response = self.client.get("/api/entries", params={"person_id": "test-candidate", "week_key": "2026-08-30"})
        entry = response.json()[0]
        for field in self.CANDIDATE_FIELDS:
            self.assertIn(field, entry, f"candidate response is missing required field: {field}")

    def test_put_response_excludes_internal_smartsheet_fields(self):
        created = self._submit().json()
        response = self.client.put(f"/api/entries/{created['id']}", json=dict(
            person_id="test-candidate", name="Test Candidate", ward="Ward 1", day="mon",
            type="Door to Door", type_display="Door to Door", notes=None,
            week_key="2026-08-30", week_label="31 Aug - 6 Sep", activity_date="2026-08-31",
            start_time="09:30", end_time="10:30", venue="Ward office",
        ))
        self.assertEqual(response.status_code, 200)
        self.assertFalse(self.INTERNAL_FIELDS & response.json().keys())

    def test_admin_review_patch_endpoint_still_returns_smartsheet_metadata(self):
        # Admin-only endpoint, gated by require_admin — must keep full fields.
        created = self._submit(type="Meeting", type_display="Meeting").json()
        token = appmod.make_admin_token()
        response = self.client.patch(
            f"/api/admin/entries/{created['id']}/smartsheet-category",
            json={"smartsheet_category": "PRESENCE"},
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["smartsheet_category"], "PRESENCE")
        self.assertIn("category_source", body)
        self.assertIn("canonical_activity", body)

    def test_other_activity_reaches_candidate_with_exact_wording_and_no_internal_terms(self):
        response = self._submit(type="Other", type_display="Community prayer event")
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertEqual(body["type_display"], "Community prayer event")
        self.assertNotEqual(body["type_display"], "Other")
        self.assertFalse(self.INTERNAL_FIELDS & body.keys())
        raw_payload = response.text
        self.assertNotIn("NEEDS_REVIEW", raw_payload)
        self.assertNotIn("smartsheet", raw_payload.lower())


@unittest.skipUnless(HAS_API_DEPS and HAS_TESTCLIENT, "FastAPI TestClient (httpx) is not installed")
class SmartSheetXlsxApiTests(unittest.TestCase):
    """SmartSheet XLSX export: real HTTP round-trip through the admin-only
    endpoint, verifying auth is enforced and the response is a genuine .xlsx
    with the exact required worksheet/column structure."""

    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("MONGO_URI", "mongodb://127.0.0.1:1")
        os.environ.setdefault("ADMIN_PIN", "1234")
        os.environ.setdefault("JWT_SECRET", "local-test-secret")
        global appmod
        import main as appmod

    def setUp(self):
        self.original_entries_col = appmod.entries_col
        self.original_roster_col = appmod.roster_col
        appmod.entries_col = FakeCollection()
        appmod.roster_col = FakeCollection()
        self.client = TestClient(appmod.app)
        self.client.__enter__()

    def tearDown(self):
        self.client.__exit__(None, None, None)
        appmod.entries_col = self.original_entries_col
        appmod.roster_col = self.original_roster_col

    def _seed(self):
        self.entries_docs = [
            entry_doc(type_display="Door to Door", week_key="2026-08-30", day="mon", start_time="09:00", end_time="10:00", venue="Area 1"),
            entry_doc(type_display="Street Meeting", week_key="2026-08-30", day="tue"),
            entry_doc(type_display="Blue Wave", week_key="2026-08-30", day="wed"),
            entry_doc(type_display="Community prayer event", type="Other", is_custom_activity=True, week_key="2026-08-30", day="thu"),
            # Legacy record: type=="Other" from before this feature existed,
            # no is_custom_activity flag -> must classify via its own text.
            entry_doc(type_display="Blue wave at entrance to town", type="Other", week_key="2026-08-30", day="fri"),
        ]
        appmod.entries_col.docs = self.entries_docs

    def test_export_xlsx_requires_admin_auth(self):
        self._seed()
        response = self.client.get("/api/admin/smartsheet/export.xlsx", params={"week_key": "2026-08-30", "category": "ALL"})
        self.assertEqual(response.status_code, 401)

    def test_export_xlsx_rejects_invalid_admin_token(self):
        self._seed()
        response = self.client.get(
            "/api/admin/smartsheet/export.xlsx",
            params={"week_key": "2026-08-30", "category": "ALL"},
            headers={"Authorization": "Bearer not-a-real-token"},
        )
        self.assertEqual(response.status_code, 401)

    def test_download_all_excel_with_valid_admin_token(self):
        self._seed()
        token = appmod.make_admin_token()
        response = self.client.get(
            "/api/admin/smartsheet/export.xlsx",
            params={"week_key": "2026-08-30", "category": "ALL"},
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("ntsikana-smartsheet-all-2026-08-30.xlsx", response.headers["content-disposition"])
        self.assertTrue(response.content.startswith(b"PK"))

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(response.content))
        self.assertEqual(wb.sheetnames, ["Canvassing", "Public-Street", "Presence"])
        # The unreviewed NEW Other activity must not appear anywhere...
        all_activities = [row[6].value for name in wb.sheetnames for row in wb[name].iter_rows(min_row=2)]
        self.assertNotIn("Community prayer event", all_activities)
        # ...but the LEGACY type=="Other" record must classify normally.
        presence_activities = [row[6].value for row in wb["Presence"].iter_rows(min_row=2)]
        self.assertIn("Blue Wave", presence_activities)

    def test_individual_category_excel_downloads_with_valid_admin_token(self):
        self._seed()
        token = appmod.make_admin_token()
        expectations = [
            ("CANVASSING", "Canvassing", "ntsikana-smartsheet-canvassing-2026-08-30.xlsx"),
            ("PUBLIC_STREET_MEETING", "Public-Street", "ntsikana-smartsheet-public-street-2026-08-30.xlsx"),
            ("PRESENCE", "Presence", "ntsikana-smartsheet-presence-2026-08-30.xlsx"),
        ]
        for category, sheet_name, filename in expectations:
            with self.subTest(category=category):
                response = self.client.get(
                    "/api/admin/smartsheet/export.xlsx",
                    params={"week_key": "2026-08-30", "category": category},
                    headers={"Authorization": f"Bearer {token}"},
                )
                self.assertEqual(response.status_code, 200)
                self.assertIn(filename, response.headers["content-disposition"])
                from openpyxl import load_workbook
                wb = load_workbook(io.BytesIO(response.content))
                self.assertEqual(wb.sheetnames, [sheet_name])

    def test_invalid_category_is_rejected(self):
        self._seed()
        token = appmod.make_admin_token()
        response = self.client.get(
            "/api/admin/smartsheet/export.xlsx",
            params={"week_key": "2026-08-30", "category": "BOGUS"},
            headers={"Authorization": f"Bearer {token}"},
        )
        self.assertEqual(response.status_code, 400)


class AsyncCursor:
    def __init__(self, docs):
        self.docs = [copy.deepcopy(doc) for doc in docs]

    def __aiter__(self):
        self._iter = iter(self.docs)
        return self

    async def __anext__(self):
        try:
            return next(self._iter)
        except StopIteration:
            raise StopAsyncIteration


class FakeCollection:
    def __init__(self):
        self.docs = []

    async def create_index(self, *args, **kwargs):
        return "idx"

    def find(self, query=None, projection=None):
        query = query or {}
        docs = [doc for doc in self.docs if matches(doc, query)]
        if projection:
            docs = [project(doc, projection) for doc in docs]
        return AsyncCursor(docs)

    async def find_one(self, query):
        for doc in self.docs:
            if matches(doc, query):
                return copy.deepcopy(doc)
        return None

    async def insert_one(self, doc):
        stored = copy.deepcopy(doc)
        stored["_id"] = ObjectId()
        self.docs.append(stored)
        return SimpleNamespace(inserted_id=stored["_id"])

    async def find_one_and_update(self, query, update, return_document=True):
        for doc in self.docs:
            if matches(doc, query):
                doc.update(copy.deepcopy(update.get("$set", {})))
                return copy.deepcopy(doc)
        return None

    async def delete_one(self, query):
        before = len(self.docs)
        self.docs = [doc for doc in self.docs if not matches(doc, query)]
        return SimpleNamespace(deleted_count=before - len(self.docs))


def matches(doc, query):
    return all(doc.get(key) == value for key, value in query.items())


def project(doc, projection):
    projected = {}
    include_keys = [key for key, value in projection.items() if value]
    if include_keys:
        for key in include_keys:
            if key in doc:
                projected[key] = doc[key]
    else:
        projected = copy.deepcopy(doc)
    if projection.get("_id") == 0:
        projected.pop("_id", None)
    return projected


def entry_doc(
    _id=None,
    type_display="Door to Door",
    type=None,
    week_key="2026-08-30",
    day="mon",
    start_time=None,
    end_time=None,
    venue=None,
    is_custom_activity=None,
):
    doc = {
        "_id": _id or ObjectId(),
        "person_id": "test-candidate",
        "name": "Test Candidate",
        "ward": "Ward 1",
        "day": day,
        "type": type or type_display,
        "type_display": type_display,
        "notes": None,
        "week_key": week_key,
        "week_label": "31 Aug - 6 Sep",
        "submitted_at": "2026-09-01T10:00:00+00:00",
    }
    if start_time is not None:
        doc["start_time"] = start_time
    if end_time is not None:
        doc["end_time"] = end_time
    if venue is not None:
        doc["venue"] = venue
    if is_custom_activity is not None:
        doc["is_custom_activity"] = is_custom_activity
    return doc


async def streaming_body(response):
    chunks = []
    async for chunk in response.body_iterator:
        chunks.append(chunk if isinstance(chunk, bytes) else chunk.encode())
    return b"".join(chunks)


def csv_rows(payload):
    return list(csv.reader(io.StringIO(payload.decode("utf-8-sig"))))


if __name__ == "__main__":
    unittest.main()
