import asyncio
import copy
import os
import unittest
from types import SimpleNamespace

# official_capture.py imports openpyxl lazily (only inside
# official_capture_xlsx_bytes), so this top-level import works even in an
# environment with no Excel library installed — every non-xlsx test class
# below runs for real, not just "skipped", in that environment.
import official_capture as oc

try:
    from openpyxl import load_workbook

    HAS_OPENPYXL = True
except ModuleNotFoundError:
    load_workbook = None
    HAS_OPENPYXL = False

try:
    import fastapi  # noqa: F401
    from bson import ObjectId
    from fastapi import HTTPException

    HAS_API_DEPS = True
except ModuleNotFoundError:
    ObjectId = None
    HTTPException = Exception
    HAS_API_DEPS = False

try:
    from fastapi.testclient import TestClient  # requires httpx

    HAS_TESTCLIENT = True
except ImportError:
    TestClient = None
    HAS_TESTCLIENT = False


def load_wb(payload: bytes):
    return load_workbook(__import__("io").BytesIO(payload))


# ---------------------------------------------------------------------------
# Pure official_capture.py tests — no FastAPI/Mongo needed at all.
# ---------------------------------------------------------------------------
class OfficialTypeMappingTests(unittest.TestCase):
    def test_confident_mapping_gives_expected_suggestion(self):
        cases = {
            "Door to Door": "In-person Canvassing / Door-to-door",
            "Info Table": "Info Table",
            "Info table : canvassing": "Info Table",
            "Telecanvassing": "Tele Canvassing",
            "House Meeting": "House meeting",
            "Public Meeting": "Public meeting",
            "Clean up": "Clean-up Event",
            "Oversight": "Oversight Visit",
            "Stakeholder meeting": "Stakeholder Meeting",
            "Hoot or Blue wave": "Blue Wave / Robot blitz",
            "Blue Wave": "Blue Wave / Robot blitz",
            "Motorcade": "Cavalcade / Carcade / Motorcade",
            "March": "March",
            "Picket": "Protest / Picket",
            "Rally": "Rally",
            "Religious Forum Address": "Religious Forum Address",
            "Poster fighting": "Poster fighting",
            "Leaflet Distribution": "Leaflet distribution",
            "Care Event": "Care event (Oppit)",
        }
        for ward_tracker_type, expected in cases.items():
            with self.subTest(ward_tracker_type=ward_tracker_type):
                self.assertEqual(
                    oc.suggested_official_type({"type_display": ward_tracker_type}), expected
                )

    def test_confident_mapping_is_case_and_punctuation_insensitive(self):
        self.assertEqual(
            oc.suggested_official_type({"type_display": "door to door"}),
            "In-person Canvassing / Door-to-door",
        )
        self.assertEqual(
            oc.suggested_official_type({"type_display": "  DOOR TO DOOR  "}),
            "In-person Canvassing / Door-to-door",
        )

    def test_ambiguous_mapping_does_not_silently_map(self):
        # "Street Meeting" is a real, audited Ward Tracker activity (see
        # smartsheet_reporting.CANONICAL_ACTIVITY_CATEGORY) but is NOT one of
        # the confidently-mapped official types — it must stay unmapped here.
        self.assertIsNone(oc.suggested_official_type({"type_display": "Street Meeting"}))
        self.assertIsNone(oc.suggested_official_type({"type_display": "Sports day"}))
        self.assertIsNone(oc.suggested_official_type({"type_display": "Social Media"}))

    def test_unmapped_type_does_not_silently_map(self):
        self.assertIsNone(oc.suggested_official_type({"type_display": "Completely Made Up Activity"}))
        self.assertIsNone(oc.suggested_official_type({"type_display": ""}))
        self.assertIsNone(oc.suggested_official_type({}))

    def test_resolve_prefers_override_over_suggestion(self):
        doc = {"type_display": "Door to Door", "official_activity_type": "Rally"}
        self.assertEqual(oc.resolve_official_activity_type(doc), "Rally")

    def test_resolve_falls_back_to_suggestion_when_no_override(self):
        doc = {"type_display": "Door to Door"}
        self.assertEqual(
            oc.resolve_official_activity_type(doc), "In-person Canvassing / Door-to-door"
        )

    def test_validate_official_activity_type_accepts_whitelisted_value(self):
        self.assertEqual(oc.validate_official_activity_type("Rally"), "Rally")

    def test_validate_official_activity_type_rejects_arbitrary_string(self):
        with self.assertRaises(ValueError):
            oc.validate_official_activity_type("Something I Just Invented")

    def test_full_official_type_list_has_exactly_46_types(self):
        self.assertEqual(len(oc.OFFICIAL_ACTIVITY_TYPES), 46)
        self.assertEqual(len(oc.OFFICIAL_ACTIVITY_TYPES), len(set(oc.OFFICIAL_ACTIVITY_TYPES)))

    def test_confident_mapping_targets_are_a_subset_of_the_full_list(self):
        confident_targets = set(oc._CONFIDENT_MAP_RAW.values())
        self.assertEqual(len(confident_targets), 17)
        self.assertTrue(confident_targets.issubset(set(oc.OFFICIAL_ACTIVITY_TYPES)))

    def test_admin_can_validate_a_full_list_type_that_is_not_a_confident_target(self):
        # "Billboard" is on the full 46-type list but is never an automatic
        # suggestion for anything — proves the dropdown/validator is not
        # secretly still limited to the 17 confident-mapping targets.
        self.assertNotIn("Billboard", oc._CONFIDENT_MAP_RAW.values())
        self.assertEqual(oc.validate_official_activity_type("Billboard"), "Billboard")

    def test_ambiguous_activity_can_be_manually_resolved_to_any_full_list_type(self):
        # "Street Meeting" gets no automatic suggestion, but the coordinator
        # must be able to confirm it as literally any of the 46 known types,
        # not just one of the 17 confident-mapping targets.
        doc = {"type_display": "Street Meeting"}
        self.assertIsNone(oc.suggested_official_type(doc))
        doc["official_activity_type"] = oc.validate_official_activity_type("Community Crime Patrol")
        self.assertEqual(oc.resolve_official_activity_type(doc), "Community Crime Patrol")

    def test_original_ward_tracker_activity_type_is_never_touched(self):
        doc = {"type_display": "Door to Door", "type": "Door to Door"}
        oc.resolve_official_activity_type(doc)
        self.assertEqual(doc["type_display"], "Door to Door")
        self.assertEqual(doc["type"], "Door to Door")


class CaptureStatusHelperTests(unittest.TestCase):
    def test_historical_entry_missing_capture_status_resolves_safely(self):
        self.assertEqual(oc.resolve_capture_status({}), oc.AWAITING_CAPTURE)

    def test_unexpected_stored_value_also_resolves_safely(self):
        self.assertEqual(oc.resolve_capture_status({"capture_status": "bogus"}), oc.AWAITING_CAPTURE)

    def test_stored_captured_resolves_to_captured(self):
        self.assertEqual(oc.resolve_capture_status({"capture_status": "captured"}), oc.CAPTURED)

    def test_validate_capture_status_rejects_invalid_value(self):
        with self.assertRaises(ValueError):
            oc.validate_capture_status("processing")

    def test_validate_capture_status_accepts_known_values(self):
        self.assertEqual(oc.validate_capture_status("awaiting_capture"), "awaiting_capture")
        self.assertEqual(oc.validate_capture_status("captured"), "captured")


class AugmentFilterSortTests(unittest.TestCase):
    def _doc(self, **overrides):
        doc = {
            "id": "e1",
            "person_id": "test-candidate",
            "name": "Test Candidate",
            "ward": "Ward 1",
            "activity_date": "2026-09-10",
            "start_time": "16:00",
            "end_time": "18:00",
            "type": "Door to Door",
            "type_display": "Door to Door",
            "venue": "Community Hall",
            "campaign_id": None,
        }
        doc.update(overrides)
        return doc

    def test_campaign_linked_activity_displays_correct_campaign_name(self):
        row = oc.augment_entry(self._doc(campaign_id="c1"), "September Canvassing")
        self.assertEqual(row["campaign_name"], "September Canvassing")
        self.assertEqual(row["campaign_id"], "c1")

    def test_ordinary_activity_works_without_campaign(self):
        row = oc.augment_entry(self._doc(campaign_id=None), None)
        self.assertIsNone(row["campaign_name"])
        self.assertIsNone(row["campaign_id"])

    def test_augment_marks_confirmed_vs_suggested_vs_unmapped(self):
        confirmed = oc.augment_entry(self._doc(official_activity_type="Rally"), None)
        self.assertEqual(confirmed["official_activity_type_source"], "confirmed")
        self.assertEqual(confirmed["official_activity_type"], "Rally")

        suggested = oc.augment_entry(self._doc(type_display="Door to Door"), None)
        self.assertEqual(suggested["official_activity_type_source"], "suggested")

        unmapped = oc.augment_entry(self._doc(type_display="Street Meeting"), None)
        self.assertEqual(unmapped["official_activity_type_source"], "unmapped")
        self.assertIsNone(unmapped["official_activity_type"])

    def test_compute_counts(self):
        rows = [
            oc.augment_entry(self._doc(id="a"), None),
            oc.augment_entry(self._doc(id="b", capture_status="captured"), None),
            oc.augment_entry(self._doc(id="c", capture_status="captured"), None),
        ]
        counts = oc.compute_counts(rows)
        self.assertEqual(counts, {"awaiting_capture": 1, "captured": 2, "total": 3})

    def test_status_filter_awaiting_captured_all(self):
        rows = [
            oc.augment_entry(self._doc(id="a"), None),
            oc.augment_entry(self._doc(id="b", capture_status="captured"), None),
        ]
        self.assertEqual([r["id"] for r in oc.filter_entries(rows, status="awaiting_capture")], ["a"])
        self.assertEqual([r["id"] for r in oc.filter_entries(rows, status="captured")], ["b"])
        self.assertEqual(len(oc.filter_entries(rows, status="all")), 2)
        self.assertEqual(len(oc.filter_entries(rows)), 2)

    def test_needs_confirmation_filter_value(self):
        rows = [
            oc.augment_entry(self._doc(id="a", type_display="Street Meeting"), None),
            oc.augment_entry(self._doc(id="b", type_display="Door to Door"), None),
        ]
        matched = oc.filter_entries(rows, official_activity_type=oc.NEEDS_CONFIRMATION_FILTER_VALUE)
        self.assertEqual([r["id"] for r in matched], ["a"])

    def test_date_range_filter(self):
        rows = [
            oc.augment_entry(self._doc(id="a", activity_date="2026-09-01"), None),
            oc.augment_entry(self._doc(id="b", activity_date="2026-09-15"), None),
            oc.augment_entry(self._doc(id="c", activity_date="2026-09-30"), None),
        ]
        matched = oc.filter_entries(rows, date_from="2026-09-10", date_to="2026-09-20")
        self.assertEqual([r["id"] for r in matched], ["b"])

    def test_sort_oldest_first(self):
        rows = [
            oc.augment_entry(self._doc(id="later", activity_date="2026-09-20"), None),
            oc.augment_entry(self._doc(id="earlier", activity_date="2026-09-01"), None),
        ]
        ordered = oc.sort_oldest_first(rows)
        self.assertEqual([r["id"] for r in ordered], ["earlier", "later"])


@unittest.skipUnless(HAS_OPENPYXL, "openpyxl is not installed")
class OfficialCaptureXlsxTests(unittest.TestCase):
    def _row(self, **overrides):
        row = {
            "activity_date": "2026-09-08",
            "start_time": "16:00",
            "end_time": "18:00",
            "name": "Example Candidate",
            "ward": "Ward 13",
            "campaign_name": "September Canvassing",
            "type_display": "Door to Door",
            "official_activity_type": "In-person Canvassing / Door-to-door",
            "venue": "Mlungisi Community Hall",
            "capture_status": "awaiting_capture",
            "captured_at": None,
        }
        row.update(overrides)
        return row

    def test_correct_columns(self):
        wb = load_wb(oc.official_capture_xlsx_bytes([self._row()]))
        ws = wb.active
        header = [c.value for c in ws[1]]
        self.assertEqual(header, oc.OFFICIAL_CAPTURE_HEADERS)

    def test_correct_values(self):
        wb = load_wb(oc.official_capture_xlsx_bytes([self._row()]))
        ws = wb.active
        row = [c.value for c in ws[2]]
        self.assertEqual(row, [
            "2026-09-08", "16:00", "18:00", "Example Candidate", "Ward 13",
            "September Canvassing", "Door to Door", "In-person Canvassing / Door-to-door",
            "Mlungisi Community Hall", "Awaiting Capture", None,
        ])

    def test_campaign_name_included_and_dash_when_absent(self):
        wb = load_wb(oc.official_capture_xlsx_bytes([self._row(campaign_name=None)]))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=6).value, "—")

    def test_capture_status_included_and_captured_at_populated(self):
        wb = load_wb(oc.official_capture_xlsx_bytes([
            self._row(capture_status="captured", captured_at="2026-09-09T10:00:00+00:00")
        ]))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=10).value, "Captured")
        self.assertEqual(ws.cell(row=2, column=11).value, "2026-09-09T10:00:00+00:00")

    def test_needs_confirmation_label_when_unmapped(self):
        wb = load_wb(oc.official_capture_xlsx_bytes([self._row(official_activity_type=None)]))
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=8).value, oc.NEEDS_CONFIRMATION_LABEL)

    def test_spreadsheet_injection_protection_remains_effective(self):
        malicious = self._row(
            name="=cmd()", ward="+HYPERLINK(1)", campaign_name="-2+3",
            type_display="@SUM(1)", venue="=1+1",
        )
        wb = load_wb(oc.official_capture_xlsx_bytes([malicious]))
        ws = wb.active
        row = [c.value for c in ws[2]]
        self.assertEqual(row[3], "'=cmd()")
        self.assertEqual(row[4], "'+HYPERLINK(1)")
        self.assertEqual(row[5], "'-2+3")
        self.assertEqual(row[6], "'@SUM(1)")
        self.assertEqual(row[8], "'=1+1")


# ---------------------------------------------------------------------------
# FastAPI-integration tests (endpoint + document lifecycle) — mirrors the
# FakeCollection conventions in tests/test_campaigns.py and
# tests/test_api_smartsheet.py.
# ---------------------------------------------------------------------------
@unittest.skipUnless(HAS_API_DEPS, "FastAPI dependencies are not installed")
class CaptureLifecycleTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("MONGO_URI", "mongodb://127.0.0.1:1")
        os.environ.setdefault("ADMIN_PIN", "1234")
        os.environ.setdefault("JWT_SECRET", "local-test-secret")
        global appmod
        import main as appmod

    def setUp(self):
        self.original_entries_col = appmod.entries_col
        self.original_roster_col = appmod.roster_col
        self.original_campaigns_col = appmod.campaigns_col
        self.entries = FakeCollection()
        self.roster = FakeCollection()
        self.campaigns = FakeCollection()
        self.roster.docs = [
            {"_id": ObjectId(), "name": "Test Candidate", "ward": "Ward 1", "name_slug": "test-candidate"},
        ]
        appmod.entries_col = self.entries
        appmod.roster_col = self.roster
        appmod.campaigns_col = self.campaigns

    def tearDown(self):
        appmod.entries_col = self.original_entries_col
        appmod.roster_col = self.original_roster_col
        appmod.campaigns_col = self.original_campaigns_col

    def _entry_body(self, **overrides):
        kwargs = dict(
            person_id="test-candidate", name="Test Candidate", ward="Ward 1",
            day="mon", type="Door to Door", type_display="Door to Door",
            week_key="2026-08-30", week_label="31 Aug - 6 Sep",
            venue="Community Hall",
        )
        kwargs.update(overrides)
        return appmod.EntryIn(**kwargs)

    # ---- capture status: creation ----

    def test_new_activity_resolves_to_awaiting_capture(self):
        created = asyncio.run(appmod.create_entry(self._entry_body()))
        self.assertEqual(self.entries.docs[0]["capture_status"], "awaiting_capture")

    def test_historical_entry_missing_capture_status_resolves_safely_via_admin_all(self):
        self.entries.docs = [entry_doc()]  # no capture_status key at all
        self.assertNotIn("capture_status", self.entries.docs[0])
        result = asyncio.run(appmod.admin_all(True))
        self.assertEqual(result["entries"][0]["capture_status"], "awaiting_capture")
        # Never persisted back.
        self.assertNotIn("capture_status", self.entries.docs[0])

    def test_campaign_activity_creation_resolves_to_awaiting_capture(self):
        campaign_id = str(ObjectId())
        self.campaigns.docs = [{
            "_id": ObjectId(campaign_id), "person_id": "test-candidate", "name": "Sept Drive",
            "start_date": "2026-09-01", "end_date": "2026-09-30", "archived_at": None,
        }]
        body = appmod.CampaignActivityIn(
            person_id="test-candidate", activity_date="2026-09-10",
            type="Door to Door", type_display="Door to Door", venue="Hall",
        )
        asyncio.run(appmod.create_campaign_activity(campaign_id, body))
        self.assertEqual(self.entries.docs[0]["capture_status"], "awaiting_capture")

    def test_repeat_activities_all_resolve_to_awaiting_capture(self):
        campaign_id = str(ObjectId())
        self.campaigns.docs = [{
            "_id": ObjectId(campaign_id), "person_id": "test-candidate", "name": "Sept Drive",
            "start_date": "2026-09-01", "end_date": "2026-09-30", "archived_at": None,
        }]
        body = appmod.CampaignActivityRepeatIn(
            person_id="test-candidate", weekday="mon",
            first_occurrence_date="2026-09-07", until="2026-09-21",
            type="Door to Door", type_display="Door to Door", venue="Hall",
        )
        asyncio.run(appmod.create_campaign_activity_repeat(campaign_id, body))
        self.assertEqual(len(self.entries.docs), 3)
        self.assertTrue(all(d["capture_status"] == "awaiting_capture" for d in self.entries.docs))

    # ---- capture status: mark captured / undo ----

    def test_mark_captured_persists_captured_and_populates_captured_at(self):
        created = asyncio.run(appmod.create_entry(self._entry_body()))
        result = asyncio.run(appmod.update_entry_capture(
            created["id"], appmod.CaptureUpdateIn(capture_status="captured"), True
        ))
        self.assertEqual(result["capture_status"], "captured")
        self.assertIsNotNone(result["captured_at"])
        self.assertEqual(self.entries.docs[0]["capture_status"], "captured")
        self.assertIsNotNone(self.entries.docs[0]["captured_at"])

    def test_captured_at_uses_server_side_time_never_client_supplied(self):
        created = asyncio.run(appmod.create_entry(self._entry_body()))
        # CaptureUpdateIn has no captured_at field at all — an attempt to
        # smuggle one in is structurally impossible, not merely overwritten.
        self.assertFalse(hasattr(appmod.CaptureUpdateIn(capture_status="captured"), "captured_at"))
        result = asyncio.run(appmod.update_entry_capture(
            created["id"], appmod.CaptureUpdateIn(capture_status="captured"), True
        ))
        from datetime import datetime, timezone
        captured_at = datetime.fromisoformat(result["captured_at"])
        self.assertLess((datetime.now(timezone.utc) - captured_at).total_seconds(), 30)

    def test_undo_returns_to_awaiting_capture_and_clears_captured_at(self):
        created = asyncio.run(appmod.create_entry(self._entry_body()))
        asyncio.run(appmod.update_entry_capture(
            created["id"], appmod.CaptureUpdateIn(capture_status="captured"), True
        ))
        result = asyncio.run(appmod.update_entry_capture(
            created["id"], appmod.CaptureUpdateIn(capture_status="awaiting_capture"), True
        ))
        self.assertEqual(result["capture_status"], "awaiting_capture")
        self.assertIsNone(result["captured_at"])
        self.assertIsNone(self.entries.docs[0]["captured_at"])

    def test_invalid_capture_status_rejected(self):
        created = asyncio.run(appmod.create_entry(self._entry_body()))
        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.update_entry_capture(
                created["id"], appmod.CaptureUpdateIn(capture_status="processing"), True
            ))
        self.assertEqual(exc.exception.status_code, 400)

    def test_editing_a_captured_activity_does_not_reset_capture_status(self):
        created = asyncio.run(appmod.create_entry(self._entry_body()))
        asyncio.run(appmod.update_entry_capture(
            created["id"], appmod.CaptureUpdateIn(capture_status="captured"), True
        ))
        captured_at_before = self.entries.docs[0]["captured_at"]
        # A normal candidate edit (e.g. changing the venue) must never touch
        # capture_status/captured_at.
        asyncio.run(appmod.update_entry(created["id"], self._entry_body(venue="New Venue")))
        self.assertEqual(self.entries.docs[0]["capture_status"], "captured")
        self.assertEqual(self.entries.docs[0]["captured_at"], captured_at_before)

    def test_nothing_to_update_rejected(self):
        created = asyncio.run(appmod.create_entry(self._entry_body()))
        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.update_entry_capture(created["id"], appmod.CaptureUpdateIn(), True))
        self.assertEqual(exc.exception.status_code, 400)

    def test_capture_update_missing_entry_404s(self):
        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.update_entry_capture(
                str(ObjectId()), appmod.CaptureUpdateIn(capture_status="captured"), True
            ))
        self.assertEqual(exc.exception.status_code, 404)

    # ---- official type: admin override ----

    def test_admin_override_persists_and_wins_over_suggestion(self):
        created = asyncio.run(appmod.create_entry(self._entry_body(type="Door to Door", type_display="Door to Door")))
        result = asyncio.run(appmod.update_entry_capture(
            created["id"], appmod.CaptureUpdateIn(official_activity_type="Rally"), True
        ))
        self.assertEqual(result["official_activity_type"], "Rally")
        self.assertEqual(self.entries.docs[0]["official_activity_type"], "Rally")
        # Original Ward Tracker activity type is untouched.
        self.assertEqual(self.entries.docs[0]["type_display"], "Door to Door")
        self.assertEqual(self.entries.docs[0]["type"], "Door to Door")

    def test_ambiguous_activity_can_be_confirmed_with_any_full_list_type(self):
        # "Street Meeting" has no confident automatic suggestion, but the
        # coordinator must still be able to confirm it as any of the full
        # 46-type list — not just one of the 17 confident-mapping targets.
        created = asyncio.run(appmod.create_entry(
            self._entry_body(type="Street Meeting", type_display="Street Meeting")
        ))
        result = asyncio.run(appmod.update_entry_capture(
            created["id"], appmod.CaptureUpdateIn(official_activity_type="Community Crime Patrol"), True
        ))
        self.assertEqual(result["official_activity_type"], "Community Crime Patrol")
        self.assertEqual(self.entries.docs[0]["type_display"], "Street Meeting")

    def test_invalid_official_type_rejected(self):
        created = asyncio.run(appmod.create_entry(self._entry_body()))
        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.update_entry_capture(
                created["id"], appmod.CaptureUpdateIn(official_activity_type="Not A Real Type"), True
            ))
        self.assertEqual(exc.exception.status_code, 400)
        self.assertNotIn("official_activity_type", self.entries.docs[0])

    # ---- campaign name resolution ----

    def test_official_capture_list_resolves_campaign_name(self):
        campaign_id = str(ObjectId())
        self.campaigns.docs = [{
            "_id": ObjectId(campaign_id), "person_id": "test-candidate", "name": "Sept Drive",
            "start_date": "2026-09-01", "end_date": "2026-09-30", "archived_at": None,
        }]
        self.entries.docs = [entry_doc(campaign_id=campaign_id)]
        result = asyncio.run(appmod.admin_official_capture(True))
        self.assertEqual(result["entries"][0]["campaign_name"], "Sept Drive")

    def test_official_capture_list_shows_dash_equivalent_none_without_campaign(self):
        self.entries.docs = [entry_doc()]
        result = asyncio.run(appmod.admin_official_capture(True))
        self.assertIsNone(result["entries"][0]["campaign_name"])
        self.assertIsNone(result["entries"][0]["campaign_id"])

    # ---- ordering / counts ----

    def test_official_capture_list_is_oldest_first_by_default(self):
        self.entries.docs = [
            entry_doc(_id=ObjectId(), activity_date_override="2026-09-20"),
            entry_doc(_id=ObjectId(), activity_date_override="2026-09-01"),
        ]
        result = asyncio.run(appmod.admin_official_capture(True))
        dates = [e["activity_date"] for e in result["entries"]]
        self.assertEqual(dates, sorted(dates))

    def test_official_capture_counts_reflect_full_dataset(self):
        self.entries.docs = [
            entry_doc(_id=ObjectId(), capture_status="captured"),
            entry_doc(_id=ObjectId()),
            entry_doc(_id=ObjectId()),
        ]
        result = asyncio.run(appmod.admin_official_capture(True))
        self.assertEqual(result["counts"], {"awaiting_capture": 2, "captured": 1, "total": 3})

    # ---- regression: normal / campaign / recurrence paths unaffected ----

    def test_normal_candidate_activity_submission_still_works(self):
        result = asyncio.run(appmod.create_entry(self._entry_body()))
        self.assertEqual(result["type_display"], "Door to Door")
        self.assertEqual(len(self.entries.docs), 1)

    def test_archived_campaign_guard_still_works(self):
        campaign_id = str(ObjectId())
        self.campaigns.docs = [{
            "_id": ObjectId(campaign_id), "person_id": "test-candidate", "name": "Sept Drive",
            "start_date": "2026-09-01", "end_date": "2026-09-30",
            "archived_at": "2026-09-05T00:00:00+00:00",
        }]
        body = appmod.CampaignActivityIn(
            person_id="test-candidate", activity_date="2026-09-10",
            type="Door to Door", type_display="Door to Door", venue="Hall",
        )
        with self.assertRaises(HTTPException) as exc:
            asyncio.run(appmod.create_campaign_activity(campaign_id, body))
        self.assertEqual(exc.exception.status_code, 409)


@unittest.skipUnless(HAS_TESTCLIENT, "FastAPI TestClient (httpx) is not installed")
class CaptureApiHttpTests(unittest.TestCase):
    """Real HTTP round-trip — proves candidate responses stay clean and the
    admin endpoint truly requires the admin token, not just that the Python
    function accepts a hardcoded True."""

    @classmethod
    def setUpClass(cls):
        os.environ.setdefault("MONGO_URI", "mongodb://127.0.0.1:1")
        os.environ.setdefault("ADMIN_PIN", "1234")
        os.environ.setdefault("JWT_SECRET", "local-test-secret")
        global appmod
        import main as appmod

    def setUp(self):
        self.original_entries_col = appmod.entries_col
        self.original_roster_col = appmod.roster_col
        self.original_campaigns_col = appmod.campaigns_col
        appmod.entries_col = FakeCollection()
        appmod.roster_col = FakeCollection()
        appmod.campaigns_col = FakeCollection()
        appmod.roster_col.docs = [
            {"_id": ObjectId(), "name": "Test Candidate", "ward": "Ward 1", "name_slug": "test-candidate"},
        ]
        self.client = TestClient(appmod.app)
        self.client.__enter__()

    def tearDown(self):
        self.client.__exit__(None, None, None)
        appmod.entries_col = self.original_entries_col
        appmod.roster_col = self.original_roster_col
        appmod.campaigns_col = self.original_campaigns_col

    def test_capture_endpoint_requires_admin_authentication(self):
        res = self.client.patch(
            "/api/admin/entries/000000000000000000000000/capture",
            json={"capture_status": "captured"},
        )
        self.assertEqual(res.status_code, 401)

    def test_official_capture_list_requires_admin_authentication(self):
        res = self.client.get("/api/admin/official-capture")
        self.assertEqual(res.status_code, 401)

    def test_official_capture_export_requires_admin_authentication(self):
        res = self.client.get("/api/admin/official-capture/export.xlsx")
        self.assertEqual(res.status_code, 401)

    def test_normal_entry_submission_response_excludes_capture_fields(self):
        res = self.client.post("/api/entries", json={
            "person_id": "test candidate", "name": "Test Candidate", "ward": "Ward 1",
            "day": "mon", "type": "Door to Door", "type_display": "Door to Door",
            "week_key": "2026-08-30", "week_label": "31 Aug - 6 Sep", "venue": "Hall",
        })
        self.assertEqual(res.status_code, 200)
        body = res.json()
        for forbidden in ("capture_status", "official_activity_type", "captured_at"):
            self.assertNotIn(forbidden, body, f"candidate response must never include {forbidden}")


def entry_doc(
    _id=None,
    type_display="Door to Door",
    type=None,
    week_key="2026-08-30",
    day="mon",
    venue="Community Hall",
    campaign_id=None,
    capture_status=None,
    activity_date_override=None,
):
    doc = {
        "_id": _id or ObjectId(),
        "person_id": "test-candidate",
        "name": "Test Candidate",
        "ward": "Ward 1",
        "day": day,
        "type": type or type_display,
        "type_display": type_display,
        "notes": None,
        "week_key": week_key,
        "week_label": "31 Aug - 6 Sep",
        "venue": venue,
        "submitted_at": "2026-09-01T10:00:00+00:00",
    }
    if campaign_id is not None:
        doc["campaign_id"] = campaign_id
    if capture_status is not None:
        doc["capture_status"] = capture_status
    if activity_date_override is not None:
        doc["activity_date"] = activity_date_override
    return doc


class AsyncCursor:
    def __init__(self, docs):
        self.docs = [copy.deepcopy(doc) for doc in docs]

    def __aiter__(self):
        self._iter = iter(self.docs)
        return self

    async def __anext__(self):
        try:
            return next(self._iter)
        except StopIteration:
            raise StopAsyncIteration


class FakeCollection:
    def __init__(self):
        self.docs = []

    async def create_index(self, *args, **kwargs):
        return "idx"

    def find(self, query=None, projection=None):
        query = query or {}
        docs = [doc for doc in self.docs if matches(doc, query)]
        return AsyncCursor(docs)

    async def find_one(self, query):
        for doc in self.docs:
            if matches(doc, query):
                return copy.deepcopy(doc)
        return None

    async def insert_one(self, doc):
        stored = copy.deepcopy(doc)
        stored["_id"] = ObjectId()
        self.docs.append(stored)
        return SimpleNamespace(inserted_id=stored["_id"])

    async def find_one_and_update(self, query, update, return_document=True):
        for doc in self.docs:
            if matches(doc, query):
                doc.update(copy.deepcopy(update.get("$set", {})))
                return copy.deepcopy(doc)
        return None

    async def delete_one(self, query):
        before = len(self.docs)
        self.docs = [doc for doc in self.docs if not matches(doc, query)]
        return SimpleNamespace(deleted_count=before - len(self.docs))


def matches(doc, query):
    return all(doc.get(key) == value for key, value in query.items())


if __name__ == "__main__":
    unittest.main()
