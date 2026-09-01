import copy
import csv
import io
import unittest

from openpyxl import load_workbook

from smartsheet_reporting import (
    CANONICAL_ACTIVITY_CATEGORY,
    CANVASSING,
    CUSTOM_OTHER_TYPE,
    NEEDS_REVIEW,
    PRESENCE,
    PUBLIC_STREET_MEETING,
    SMARTSHEET_HEADERS,
    SMARTSHEET_WORKSHEET_NAMES,
    classification_for_entry,
    classify_activity_text,
    is_custom_other_entry,
    is_new_other_submission,
    review_entries,
    reporting_metadata_for_submission,
    smartsheet_csv_bytes,
    smartsheet_rows,
    smartsheet_workbook_all_categories_bytes,
    smartsheet_xlsx_bytes,
    spreadsheet_safe_text,
    summarize_smartsheet_entries,
    validate_time_range,
)


def csv_rows(payload: bytes) -> list[list[str]]:
    return list(csv.reader(io.StringIO(payload.decode("utf-8-sig"))))


def load_wb(payload: bytes):
    return load_workbook(io.BytesIO(payload))


class SmartSheetReportingTests(unittest.TestCase):
    def test_each_approved_activity_maps_to_one_destination(self):
        for label, expected_category in CANONICAL_ACTIVITY_CATEGORY.items():
            with self.subTest(label=label):
                classification = classify_activity_text(label)
                self.assertEqual(classification.category, expected_category)
                self.assertEqual(classification.canonical_activity, label)
                self.assertFalse(classification.needs_review)

    def test_required_specific_mappings(self):
        self.assertEqual(classify_activity_text("House Meeting").category, CANVASSING)
        self.assertEqual(classify_activity_text("Street Meeting").category, PUBLIC_STREET_MEETING)
        self.assertEqual(classify_activity_text("Blue Wave").category, PRESENCE)

    def test_case_insensitive_historical_matching(self):
        self.assertEqual(classify_activity_text("door to door canvassing Ward 5").category, CANVASSING)
        self.assertEqual(classify_activity_text("PUBLIC MEETING").category, PUBLIC_STREET_MEETING)
        self.assertEqual(classify_activity_text("cleaning up dump site").category, PRESENCE)
        self.assertEqual(classify_activity_text("SOUP kitchen").category, PRESENCE)

    def test_unclear_historical_text_needs_review(self):
        classification = classify_activity_text("candidate follow up")
        self.assertEqual(classification.category, NEEDS_REVIEW)
        self.assertTrue(classification.needs_review)

    def test_generic_canvassing_is_suggested_but_still_needs_review_for_exact_activity(self):
        classification = classify_activity_text("Canvassing")
        self.assertEqual(classification.category, CANVASSING)
        self.assertEqual(classification.suggested_category, CANVASSING)
        self.assertTrue(classification.needs_review)

    def test_poster_preparation_is_suggested_presence_but_still_needs_review(self):
        classification = classify_activity_text("Preparing posters with cable ties")
        self.assertEqual(classification.category, PRESENCE)
        self.assertEqual(classification.suggested_category, PRESENCE)
        self.assertIsNone(classification.canonical_activity)
        self.assertTrue(classification.needs_review)

    def test_time_validation_accepts_valid_times_and_rejects_obvious_invalid_ranges(self):
        self.assertEqual(validate_time_range("09:00", "10:30"), ("09:00", "10:30"))
        self.assertEqual(validate_time_range(None, None), (None, None))
        with self.assertRaises(ValueError):
            validate_time_range("10:00", "09:30")
        with self.assertRaises(ValueError):
            validate_time_range("9am", "10:00")

    def test_new_submission_derives_category_and_canonical_activity(self):
        metadata = reporting_metadata_for_submission({
            "type": "Door to Door",
            "type_display": "Door to Door",
        })
        self.assertEqual(metadata["smartsheet_category"], CANVASSING)
        self.assertEqual(metadata["canonical_activity"], "Door to Door")
        self.assertEqual(metadata["category_source"], "automatic")

    def test_admin_review_category_is_preserved_when_activity_text_is_unchanged(self):
        existing = {
            "type": "Other",
            "type_display": "Legacy activity",
            "smartsheet_category": PRESENCE,
            "category_source": "admin_review",
            "category_reviewed_at": "2026-09-01T10:00:00+00:00",
        }
        metadata = reporting_metadata_for_submission({
            "type": "Other",
            "type_display": "Legacy activity",
        }, existing)
        self.assertEqual(metadata["smartsheet_category"], PRESENCE)
        self.assertEqual(metadata["category_source"], "admin_review")

    def test_legacy_records_remain_readable_and_unchanged_by_classification(self):
        entries = [
            {
                "id": "1",
                "week_key": "2026-08-30",
                "day": "mon",
                "type": "Door to Door",
                "type_display": "Door to Door",
                "ward": "Ward 1",
            },
            {
                "id": "2",
                "week_key": "2026-08-30",
                "day": "tue",
                "type": "Other",
                "type_display": "Legacy unclear activity",
                "ward": "Ward 2",
            },
            {
                "id": "3",
                "week_key": "2026-08-23",
                "day": "sun",
                "type": "Blue Wave",
                "type_display": "Blue Wave",
                "ward": "Ward 3",
            },
        ]
        before = copy.deepcopy(entries)

        summary = summarize_smartsheet_entries(entries, "2026-08-30")
        review_rows = review_entries(entries)

        self.assertEqual(len(entries), 3)
        self.assertEqual(len(before), len(entries))
        self.assertEqual(before, entries)
        self.assertEqual(summary["total_historical_activities"], 3)
        self.assertEqual(summary["weekly"]["total"], 2)
        self.assertEqual(summary["needs_review"], len(review_rows))
        self.assertEqual(summary["needs_review"], 1)
        self.assertTrue(all("start_time" not in entry for entry in entries))
        self.assertTrue(all("end_time" not in entry for entry in entries))
        self.assertTrue(all("venue" not in entry for entry in entries))

    def test_smartsheet_export_columns_week_filtering_and_blank_historical_times(self):
        docs = [
            {
                "id": "1",
                "week_key": "2026-08-30",
                "day": "mon",
                "activity_date": "2026-08-31",
                "ward": "Ward 1",
                "venue": "Hall",
                "type": "Door to Door",
                "type_display": "Door to Door",
                "start_time": "09:00",
                "end_time": "10:00",
            },
            {
                "id": "2",
                "week_key": "2026-08-30",
                "day": "tue",
                "activity_date": "2026-09-01",
                "ward": "Ward 2",
                "type": "Street Meeting",
                "type_display": "Street Meeting",
            },
            {
                "id": "3",
                "week_key": "2026-08-23",
                "day": "mon",
                "activity_date": "2026-08-24",
                "ward": "Ward 3",
                "type": "Blue Wave",
                "type_display": "Blue Wave",
            },
        ]

        rows = csv_rows(smartsheet_csv_bytes(docs, "2026-08-30", CANVASSING))
        self.assertEqual(rows[0], SMARTSHEET_HEADERS)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][:7], ["2026-08-31", "09:00", "10:00", "Ntsikana Constituency", "Ward 1", "Hall", "Door to Door"])

        public_rows = csv_rows(smartsheet_csv_bytes(docs, "2026-08-30", PUBLIC_STREET_MEETING))
        self.assertEqual(public_rows[1][1], "")
        self.assertEqual(public_rows[1][2], "")
        self.assertEqual(public_rows[1][6], "Street Meeting")

    def test_no_duplicate_rows_between_category_exports(self):
        docs = [
            {"id": "1", "week_key": "2026-08-30", "day": "mon", "type_display": "Door to Door", "ward": "1"},
            {"id": "2", "week_key": "2026-08-30", "day": "tue", "type_display": "Street Meeting", "ward": "2"},
            {"id": "3", "week_key": "2026-08-30", "day": "wed", "type_display": "Blue Wave", "ward": "3"},
            {"id": "4", "week_key": "2026-08-30", "day": "thu", "type_display": "Unclear work", "ward": "4"},
        ]
        category_rows = []
        for category in (CANVASSING, PUBLIC_STREET_MEETING, PRESENCE):
            category_rows.extend(tuple(row) for row in smartsheet_rows(docs, "2026-08-30", category))

        self.assertEqual(len(category_rows), 3)
        self.assertEqual(len(set(category_rows)), 3)

        all_rows = csv_rows(smartsheet_csv_bytes(docs, "2026-08-30", "ALL"))
        self.assertIn("SMARTSHEET DESTINATION", all_rows[0])
        self.assertIn("REVIEW STATUS", all_rows[0])
        self.assertTrue(any(row[-2:] == ["Needs Review", "Needs review"] for row in all_rows[1:]))


class OtherActivityClassificationTests(unittest.TestCase):
    """Item 13: a candidate-typed "Other" activity must never be auto-guessed
    into a fixed SmartSheet category, even when the wording happens to match
    an official activity label exactly."""

    def test_is_custom_other_entry_requires_the_persisted_flag_not_bare_type(self):
        # The persisted is_custom_activity flag is authoritative...
        self.assertTrue(is_custom_other_entry({"type": CUSTOM_OTHER_TYPE, "type_display": "x", "is_custom_activity": True}))
        self.assertFalse(is_custom_other_entry({"type": CUSTOM_OTHER_TYPE, "type_display": "x", "is_custom_activity": False}))
        self.assertFalse(is_custom_other_entry({"type": "Door to Door", "type_display": "Door to Door"}))
        self.assertFalse(is_custom_other_entry({}))
        # ...and critically, a bare type=="Other" with NO flag at all (exactly
        # the shape of pre-existing legacy records that predate this feature
        # and used "Other" as a generic historical placeholder) must NOT match.
        self.assertFalse(is_custom_other_entry({"type": CUSTOM_OTHER_TYPE, "type_display": "Blue wave at entrance to town"}))

    def test_is_new_other_submission_checks_live_request_type_only(self):
        self.assertTrue(is_new_other_submission({"type": CUSTOM_OTHER_TYPE, "type_display": "Community prayer event"}))
        self.assertFalse(is_new_other_submission({"type": "Door to Door", "type_display": "Door to Door"}))
        self.assertFalse(is_new_other_submission({}))
        self.assertFalse(is_new_other_submission({"type": " other "}))  # must match exactly, not fuzzily

    def test_other_submission_with_custom_wording_needs_review(self):
        metadata = reporting_metadata_for_submission({
            "type": CUSTOM_OTHER_TYPE,
            "type_display": "Community prayer event",
        })
        self.assertEqual(metadata["smartsheet_category"], NEEDS_REVIEW)
        self.assertIsNone(metadata["canonical_activity"])
        self.assertEqual(metadata["category_source"], "automatic")
        self.assertFalse(metadata["category_reviewed"])

    def test_other_wording_matching_an_official_label_is_not_auto_classified(self):
        # The candidate deliberately used Other instead of the matching dropdown
        # entry — submission-time metadata must not silently reclassify it.
        metadata = reporting_metadata_for_submission({
            "type": CUSTOM_OTHER_TYPE,
            "type_display": "Door to Door",
        })
        self.assertEqual(metadata["smartsheet_category"], NEEDS_REVIEW)
        self.assertIsNone(metadata["canonical_activity"])
        self.assertTrue(metadata["is_custom_activity"])

        # And re-deriving classification straight from the STORED doc (as the
        # admin review list, summary, and exports all do) must agree — this
        # relies on the persisted is_custom_activity flag, not the raw type.
        classification = classification_for_entry({
            "type": CUSTOM_OTHER_TYPE,
            "type_display": "Door to Door",
            **metadata,
        })
        self.assertEqual(classification.category, NEEDS_REVIEW)
        self.assertIsNone(classification.canonical_activity)

    def test_new_other_entry_with_no_other_stored_metadata_still_needs_review(self):
        # A freshly-persisted new-Other record (is_custom_activity: True, but
        # no smartsheet_category/canonical_activity cached yet) must still
        # resolve to NEEDS_REVIEW when re-derived from the stored doc.
        classification = classification_for_entry({
            "type": CUSTOM_OTHER_TYPE,
            "type_display": "Public Meeting",
            "is_custom_activity": True,
        })
        self.assertEqual(classification.category, NEEDS_REVIEW)
        self.assertIsNone(classification.canonical_activity)
        self.assertTrue(classification.needs_review)

    def test_other_entry_appears_in_admin_review_with_exact_wording(self):
        entries = [{
            "id": "42",
            "week_key": "2026-08-30",
            "day": "wed",
            "type": CUSTOM_OTHER_TYPE,
            "type_display": "Community prayer event",
            "is_custom_activity": True,
            "name": "Nomsa Dlamini",
            "ward": "Ward 4",
            "smartsheet_category": NEEDS_REVIEW,
            "canonical_activity": None,
            "category_source": "automatic",
        }]
        rows = review_entries(entries)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["original_activity"], "Community prayer event")
        self.assertIsNone(rows[0]["suggested_category"])

    def test_other_entry_excluded_from_category_exports_until_reviewed(self):
        docs = [
            {"id": "1", "week_key": "2026-08-30", "day": "mon", "type": "Door to Door", "type_display": "Door to Door", "ward": "1"},
            {
                "id": "2", "week_key": "2026-08-30", "day": "tue",
                "type": CUSTOM_OTHER_TYPE, "type_display": "Door to Door", "is_custom_activity": True,
                "ward": "2", "smartsheet_category": NEEDS_REVIEW, "canonical_activity": None,
                "category_source": "automatic",
            },
        ]
        canvassing_rows = smartsheet_rows(docs, "2026-08-30", CANVASSING)
        self.assertEqual(len(canvassing_rows), 1)
        self.assertEqual(canvassing_rows[0][6], "Door to Door")

        all_rows = csv_rows(smartsheet_csv_bytes(docs, "2026-08-30", "ALL"))
        needs_review_rows = [r for r in all_rows[1:] if r[-2:] == ["Needs Review", "Needs review"]]
        self.assertEqual(len(needs_review_rows), 1)
        self.assertEqual(needs_review_rows[0][6], "Door to Door")

    def test_other_entry_admin_review_override_is_preserved_when_wording_unchanged(self):
        existing = {
            "type": CUSTOM_OTHER_TYPE,
            "type_display": "Community prayer event",
            "smartsheet_category": PRESENCE,
            "category_source": "admin_review",
            "category_reviewed_at": "2026-09-01T10:00:00+00:00",
        }
        metadata = reporting_metadata_for_submission({
            "type": CUSTOM_OTHER_TYPE,
            "type_display": "Community prayer event",
        }, existing)
        self.assertEqual(metadata["smartsheet_category"], PRESENCE)
        self.assertEqual(metadata["category_source"], "admin_review")

    def test_other_entry_edit_with_changed_wording_resets_to_needs_review(self):
        existing = {
            "type": CUSTOM_OTHER_TYPE,
            "type_display": "Community prayer event",
            "smartsheet_category": PRESENCE,
            "category_source": "admin_review",
            "category_reviewed_at": "2026-09-01T10:00:00+00:00",
        }
        metadata = reporting_metadata_for_submission({
            "type": CUSTOM_OTHER_TYPE,
            "type_display": "Different wording entirely",
        }, existing)
        self.assertEqual(metadata["smartsheet_category"], NEEDS_REVIEW)
        self.assertEqual(metadata["category_source"], "automatic")
        self.assertFalse(metadata["category_reviewed"])

    def test_normal_activities_unaffected_by_other_handling(self):
        # House Meeting -> Canvassing and Street Meeting -> Public/Street must
        # remain exactly as before; only the Other marker forces NEEDS_REVIEW.
        house = reporting_metadata_for_submission({"type": "House Meeting", "type_display": "House Meeting"})
        self.assertEqual(house["smartsheet_category"], CANVASSING)
        self.assertEqual(house["canonical_activity"], "House Meeting")

        street = reporting_metadata_for_submission({"type": "Street Meeting", "type_display": "Street Meeting"})
        self.assertEqual(street["smartsheet_category"], PUBLIC_STREET_MEETING)
        self.assertEqual(street["canonical_activity"], "Street Meeting")


class LegacyOtherTypeCollisionTests(unittest.TestCase):
    """Regression coverage for the production collision: "Other" was already
    used historically as a generic legacy activity type, long before Item 13
    introduced the candidate-facing Other workflow. A stored record with
    type == "Other" but no `is_custom_activity` flag is LEGACY data and must
    classify via its normal activity text and the existing historical rules —
    never forced to NEEDS_REVIEW merely because of its `type` value."""

    def test_legacy_blue_wave_record_classifies_via_historical_text(self):
        # Exact production record: type=="Other" from before this feature
        # existed, no smartsheet_category/canonical_activity/is_custom_activity
        # ever stored for it.
        legacy = {"type": "Other", "type_display": "Blue wave at entrance to town"}
        classification = classification_for_entry(legacy)
        self.assertEqual(classification.category, PRESENCE)
        self.assertEqual(classification.canonical_activity, "Blue Wave")
        self.assertFalse(classification.needs_review)

    def test_legacy_cleaning_up_record_classifies_via_historical_text(self):
        legacy = {"type": "Other", "type_display": "Cleaning up dump site."}
        classification = classification_for_entry(legacy)
        self.assertEqual(classification.category, PRESENCE)
        self.assertEqual(classification.canonical_activity, "Clean up")
        self.assertFalse(classification.needs_review)

    def test_legacy_other_record_with_ambiguous_text_still_needs_review(self):
        # Legacy records with genuinely ambiguous text must still end up
        # NEEDS_REVIEW — via normal classify_activity_text, not the marker.
        legacy = {"type": "Other", "type_display": "Funeral service at Catholic Church"}
        classification = classification_for_entry(legacy)
        self.assertEqual(classification.category, NEEDS_REVIEW)
        self.assertTrue(classification.needs_review)

    def test_legacy_other_records_appear_in_the_correct_category_export(self):
        docs = [
            {
                "id": "legacy-1", "week_key": "2026-08-30",
                "day": "mon", "activity_date": "2026-08-31", "ward": "Ward 1",
                "type": "Other", "type_display": "Blue wave at entrance to town",
            },
            {
                "id": "legacy-2", "week_key": "2026-08-30",
                "day": "tue", "activity_date": "2026-09-01", "ward": "Ward 2",
                "type": "Other", "type_display": "Cleaning up dump site.",
            },
        ]
        presence_rows = smartsheet_rows(docs, "2026-08-30", PRESENCE)
        self.assertEqual(len(presence_rows), 2)
        activities = {row[6] for row in presence_rows}
        self.assertEqual(activities, {"Blue Wave", "Clean up"})

        # And must NOT leak into Canvassing/Public-Street.
        self.assertEqual(smartsheet_rows(docs, "2026-08-30", CANVASSING), [])
        self.assertEqual(smartsheet_rows(docs, "2026-08-30", PUBLIC_STREET_MEETING), [])

    def test_legacy_other_records_appear_in_presence_worksheet(self):
        docs = [
            {
                "id": "legacy-1", "week_key": "2026-08-30", "day": "mon",
                "activity_date": "2026-08-31", "ward": "Ward 1",
                "type": "Other", "type_display": "Blue wave at entrance to town",
            },
        ]
        wb = load_wb(smartsheet_workbook_all_categories_bytes(docs, "2026-08-30"))
        presence_activities = [row[6].value for row in wb["Presence"].iter_rows(min_row=2)]
        self.assertEqual(presence_activities, ["Blue Wave"])
        self.assertEqual(list(wb["Canvassing"].iter_rows(min_row=2)), [])
        self.assertEqual(list(wb["Public-Street"].iter_rows(min_row=2)), [])

    def test_new_other_submission_with_same_wording_still_needs_review(self):
        # The critical distinction: identical wording, but submitted THROUGH
        # the new Other workflow (is_new_other_submission true at submission
        # time) must NOT get the legacy treatment.
        for wording in ["Blue wave at entrance to town", "Door to Door", "Cleaning up dump site."]:
            with self.subTest(wording=wording):
                metadata = reporting_metadata_for_submission({
                    "type": "Other",
                    "type_display": wording,
                })
                self.assertEqual(metadata["smartsheet_category"], NEEDS_REVIEW)
                self.assertIsNone(metadata["canonical_activity"])
                self.assertTrue(metadata["is_custom_activity"])

                stored = {"type": "Other", "type_display": wording, **metadata}
                classification = classification_for_entry(stored)
                self.assertEqual(classification.category, NEEDS_REVIEW)
                self.assertIsNone(classification.canonical_activity)

    def test_new_other_submission_excluded_from_exports_even_with_legacy_matching_text(self):
        docs = [
            # Legacy record: same wording as the new submission below, but no
            # is_custom_activity flag -> classifies normally (Presence/Blue Wave).
            {
                "id": "legacy-1", "week_key": "2026-08-30", "day": "mon",
                "activity_date": "2026-08-31", "ward": "Ward 1",
                "type": "Other", "type_display": "Blue wave at entrance to town",
            },
            # New Other submission with the identical wording -> NEEDS_REVIEW.
            {
                "id": "new-1", "week_key": "2026-08-30", "day": "tue",
                "activity_date": "2026-09-01", "ward": "Ward 2",
                "type": "Other", "type_display": "Blue wave at entrance to town",
                "smartsheet_category": NEEDS_REVIEW, "canonical_activity": None,
                "category_source": "automatic", "is_custom_activity": True,
            },
        ]
        presence_rows = smartsheet_rows(docs, "2026-08-30", PRESENCE)
        self.assertEqual(len(presence_rows), 1)  # only the legacy one

        all_rows = csv_rows(smartsheet_csv_bytes(docs, "2026-08-30", "ALL"))
        needs_review_rows = [r for r in all_rows[1:] if r[-2:] == ["Needs Review", "Needs review"]]
        self.assertEqual(len(needs_review_rows), 1)  # only the new submission

    def test_admin_review_override_still_works_for_legacy_other_record(self):
        # An admin can still manually recategorize a legacy Other record —
        # existing review workflow must be entirely unaffected by the fix.
        reviewed = {
            "type": "Other", "type_display": "Some ambiguous legacy text",
            "smartsheet_category": CANVASSING, "canonical_activity": None,
            "category_source": "admin_review", "category_reviewed": True,
            "category_reviewed_at": "2026-09-01T10:00:00+00:00",
        }
        classification = classification_for_entry(reviewed)
        self.assertEqual(classification.category, CANVASSING)


WEEK = "2026-08-30"


def _xlsx_fixture_docs():
    return [
        {
            "id": "1", "week_key": WEEK, "day": "mon", "activity_date": "2026-08-31",
            "ward": "Ward 1", "venue": "Town hall", "type": "Door to Door", "type_display": "Door to Door",
            "start_time": "09:00", "end_time": "10:00",
        },
        {
            "id": "2", "week_key": WEEK, "day": "mon", "activity_date": "2026-08-31",
            "ward": "Ward 1", "type": "House Meeting", "type_display": "House Meeting",
            # No start/end/venue — historical-style blank record.
        },
        {
            "id": "3", "week_key": WEEK, "day": "tue", "activity_date": "2026-09-01",
            "ward": "Ward 2", "venue": "Main road", "type": "Street Meeting", "type_display": "Street Meeting",
            "start_time": "17:00", "end_time": "18:30",
        },
        {
            "id": "4", "week_key": WEEK, "day": "wed", "activity_date": "2026-09-02",
            "ward": "Ward 3", "venue": "Community park", "type": "Blue Wave", "type_display": "Blue Wave",
            "start_time": "13:30", "end_time": "14:00",
        },
        {
            "id": "5", "week_key": WEEK, "day": "thu", "activity_date": "2026-09-03",
            "ward": "Ward 4", "type": "Meeting", "type_display": "Meeting",
            # Ambiguous free text -> NEEDS_REVIEW, never reviewed.
        },
        {
            "id": "6", "week_key": WEEK, "day": "fri", "activity_date": "2026-09-04",
            "ward": "Ward 5", "type": CUSTOM_OTHER_TYPE, "type_display": "Community prayer event",
            "is_custom_activity": True,
            # New Other submission (persisted flag set), unreviewed -> must stay excluded.
        },
        {
            "id": "7", "week_key": WEEK, "day": "sat", "activity_date": "2026-09-05",
            "ward": "Ward 6", "venue": "Sports field", "type": "Meeting", "type_display": "Meeting",
            "smartsheet_category": PRESENCE, "canonical_activity": None,
            "category_source": "admin_review", "category_reviewed": True,
            "category_reviewed_at": "2026-09-01T10:00:00+00:00",
            # Admin manually reviewed an ambiguous "Meeting" into Presence.
        },
    ]


class SmartSheetXlsxStructureTests(unittest.TestCase):
    def test_single_category_xlsx_is_valid_with_correct_shape(self):
        docs = _xlsx_fixture_docs()
        for category in (CANVASSING, PUBLIC_STREET_MEETING, PRESENCE):
            with self.subTest(category=category):
                payload = smartsheet_xlsx_bytes(docs, WEEK, category)
                self.assertTrue(payload.startswith(b"PK"), "must be a genuine .xlsx (zip) file, not CSV content")
                wb = load_wb(payload)
                self.assertEqual(wb.sheetnames, [SMARTSHEET_WORKSHEET_NAMES[category]])
                ws = wb.active
                self.assertEqual([c.value for c in ws[1]], SMARTSHEET_HEADERS)

    def test_download_all_is_valid_with_exactly_three_worksheets_in_order(self):
        payload = smartsheet_workbook_all_categories_bytes(_xlsx_fixture_docs(), WEEK)
        self.assertTrue(payload.startswith(b"PK"))
        wb = load_wb(payload)
        self.assertEqual(wb.sheetnames, ["Canvassing", "Public-Street", "Presence"])
        self.assertEqual(len(wb.sheetnames), 3)

    def test_no_needs_review_or_extra_worksheets_exist(self):
        wb = load_wb(smartsheet_workbook_all_categories_bytes(_xlsx_fixture_docs(), WEEK))
        forbidden = {"needs review", "summary", "instructions", "metadata", "sheet", "sheet1"}
        for name in wb.sheetnames:
            self.assertNotIn(name.strip().lower(), forbidden)

    def test_header_row_1_data_begins_row_2(self):
        wb = load_wb(smartsheet_xlsx_bytes(_xlsx_fixture_docs(), WEEK, CANVASSING))
        ws = wb.active
        self.assertEqual([c.value for c in ws[1]], SMARTSHEET_HEADERS)
        self.assertNotEqual(ws.cell(row=2, column=1).value, None)  # a real activity row, not blank/title

    def test_header_order_exact(self):
        self.assertEqual(SMARTSHEET_HEADERS, [
            "DATE", "TIME START", "TIME END", "CONSTITUENCY", "WARD", "VENUE",
            "ACTIVITY", "BOOST POST", "INFO GRAPHIC",
        ])

    def test_each_value_in_its_own_cell_no_merged_cells(self):
        wb = load_wb(smartsheet_workbook_all_categories_bytes(_xlsx_fixture_docs(), WEEK))
        for name in wb.sheetnames:
            with self.subTest(sheet=name):
                self.assertEqual(list(wb[name].merged_cells.ranges), [])

    def test_no_formulas_in_any_cell(self):
        wb = load_wb(smartsheet_workbook_all_categories_bytes(_xlsx_fixture_docs(), WEEK))
        for name in wb.sheetnames:
            for row in wb[name].iter_rows():
                for cell in row:
                    self.assertNotEqual(cell.data_type, "f", f"formula cell found in {name}!{cell.coordinate}")

    def test_header_row_is_bold_and_row_1_is_frozen(self):
        wb = load_wb(smartsheet_xlsx_bytes(_xlsx_fixture_docs(), WEEK, CANVASSING))
        ws = wb.active
        self.assertTrue(ws["A1"].font.bold)
        self.assertEqual(ws.freeze_panes, "A2")

    def test_blank_historical_time_and_venue_remain_blank(self):
        wb = load_wb(smartsheet_xlsx_bytes(_xlsx_fixture_docs(), WEEK, CANVASSING))
        ws = wb.active
        # Row for doc "2" (House Meeting, no time/venue) should have blank cells there.
        house_meeting_row = next(r for r in ws.iter_rows(min_row=2) if r[6].value == "House Meeting")
        self.assertFalse(house_meeting_row[1].value)  # TIME START
        self.assertFalse(house_meeting_row[2].value)  # TIME END
        self.assertFalse(house_meeting_row[5].value)  # VENUE
        self.assertFalse(house_meeting_row[7].value)  # BOOST POST always blank
        self.assertFalse(house_meeting_row[8].value)  # INFO GRAPHIC always blank


class SmartSheetXlsxClassificationTests(unittest.TestCase):
    def _activities_in(self, category):
        wb = load_wb(smartsheet_workbook_all_categories_bytes(_xlsx_fixture_docs(), WEEK))
        ws = wb[SMARTSHEET_WORKSHEET_NAMES[category]]
        return [row[6].value for row in ws.iter_rows(min_row=2)]

    def test_canvassing_worksheet_has_only_canvassing_records(self):
        self.assertCountEqual(self._activities_in(CANVASSING), ["Door to Door", "House Meeting"])

    def test_public_street_worksheet_has_only_public_street_records(self):
        self.assertCountEqual(self._activities_in(PUBLIC_STREET_MEETING), ["Street Meeting"])

    def test_presence_worksheet_has_only_presence_records(self):
        # Blue Wave (auto) + the admin-reviewed "Meeting" -> Presence override.
        self.assertCountEqual(self._activities_in(PRESENCE), ["Blue Wave", "Meeting"])

    def test_no_record_appears_in_more_than_one_worksheet(self):
        wb = load_wb(smartsheet_workbook_all_categories_bytes(_xlsx_fixture_docs(), WEEK))
        seen_dates_times = []
        for name in wb.sheetnames:
            for row in wb[name].iter_rows(min_row=2):
                key = (row[0].value, row[1].value, row[4].value, row[6].value)
                self.assertNotIn(key, seen_dates_times, f"duplicate row across worksheets: {key}")
                seen_dates_times.append(key)
        self.assertEqual(len(seen_dates_times), 5)  # 2 Canvassing + 1 Public-Street + 2 Presence

    def test_needs_review_records_excluded_from_every_worksheet(self):
        wb = load_wb(smartsheet_workbook_all_categories_bytes(_xlsx_fixture_docs(), WEEK))
        all_activities = [
            row[6].value
            for name in wb.sheetnames
            for row in wb[name].iter_rows(min_row=2)
        ]
        # Two docs share the raw text "Meeting": id 5 is unreviewed (NEEDS_REVIEW,
        # must be excluded) and id 7 was admin-reviewed into Presence (must appear
        # exactly once). If either leaked wrongly, this count would be 0 or 2.
        self.assertEqual(all_activities.count("Meeting"), 1)

    def test_other_activity_excluded_until_reviewed(self):
        wb = load_wb(smartsheet_workbook_all_categories_bytes(_xlsx_fixture_docs(), WEEK))
        all_activities = [
            row[6].value
            for name in wb.sheetnames
            for row in wb[name].iter_rows(min_row=2)
        ]
        self.assertNotIn("Community prayer event", all_activities)

    def test_admin_reviewed_override_appears_in_correct_worksheet(self):
        self.assertIn("Meeting", self._activities_in(PRESENCE))

    def test_house_meeting_still_maps_to_canvassing(self):
        self.assertIn("House Meeting", self._activities_in(CANVASSING))

    def test_street_meeting_still_maps_to_public_street(self):
        self.assertIn("Street Meeting", self._activities_in(PUBLIC_STREET_MEETING))


class SpreadsheetFormulaInjectionTests(unittest.TestCase):
    INJECTION_PAYLOADS = [
        "=1+1",
        '=HYPERLINK("http://example.com","click")',
        "+SUM(1,2)",
        "-1+2",
        "@SUM(1,2)",
    ]

    def test_spreadsheet_safe_text_prefixes_formula_trigger_characters(self):
        for payload in self.INJECTION_PAYLOADS:
            with self.subTest(payload=payload):
                safe = spreadsheet_safe_text(payload)
                self.assertTrue(safe.startswith("'"), f"expected a neutralizing prefix on: {payload}")
                self.assertEqual(safe[1:], payload, "original readable text must be preserved after the prefix")
                self.assertNotEqual(safe[0], "=")
                self.assertNotIn(safe[0], ("+", "-", "@"))

    def test_spreadsheet_safe_text_leaves_normal_text_untouched(self):
        for normal in ["Door to Door", "Ward 4", "Town hall", "", None]:
            self.assertEqual(spreadsheet_safe_text(normal), normal or "")

    def test_xlsx_export_neutralizes_formula_injection_in_activity_and_venue(self):
        docs = [{
            "id": "x1", "week_key": WEEK, "day": "mon", "activity_date": "2026-08-31",
            "ward": "@SUM(1,2)", "venue": "=1+1",
            "type": CUSTOM_OTHER_TYPE, "type_display": '=HYPERLINK("http://evil","click")',
            "smartsheet_category": CANVASSING, "canonical_activity": None,
            "category_source": "admin_review", "category_reviewed": True,
        }]
        wb = load_wb(smartsheet_xlsx_bytes(docs, WEEK, CANVASSING))
        ws = wb.active
        row = list(ws.iter_rows(min_row=2))[0]
        self.assertEqual(row[4].value, "'@SUM(1,2)")  # WARD
        self.assertEqual(row[5].value, "'=1+1")  # VENUE
        self.assertEqual(row[6].value, "'=HYPERLINK(\"http://evil\",\"click\")")  # ACTIVITY
        for cell in row:
            self.assertNotEqual(cell.data_type, "f")

    def test_csv_export_also_neutralizes_formula_injection(self):
        docs = [{
            "id": "x1", "week_key": WEEK, "day": "mon", "activity_date": "2026-08-31",
            "ward": "Ward 1", "venue": "+SUM(1,2)", "type": "Door to Door", "type_display": "Door to Door",
        }]
        rows = csv_rows(smartsheet_csv_bytes(docs, WEEK, CANVASSING))
        self.assertEqual(rows[1][5], "'+SUM(1,2)")

    def test_export_sanitization_does_not_mutate_the_source_document(self):
        doc = {
            "id": "x1", "week_key": WEEK, "day": "mon", "activity_date": "2026-08-31",
            "ward": "Ward 1", "venue": "=1+1", "type": "Door to Door", "type_display": "Door to Door",
        }
        before = copy.deepcopy(doc)
        smartsheet_xlsx_bytes([doc], WEEK, CANVASSING)
        smartsheet_csv_bytes([doc], WEEK, CANVASSING)
        self.assertEqual(doc, before, "the original database record must never be modified by export")

    def test_readable_text_still_understandable_to_admin(self):
        # The neutralized value must still clearly convey the original wording,
        # not be blanked out or replaced with a placeholder.
        safe = spreadsheet_safe_text("=Door to Door (urgent)")
        self.assertIn("Door to Door (urgent)", safe)


if __name__ == "__main__":
    unittest.main()
